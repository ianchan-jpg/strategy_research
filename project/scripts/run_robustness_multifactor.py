"""稳健性与多因子增量检验。

本阶段不新增或筛选股票因子，不做最终打分、机器学习或 PCA。
只使用用户指定的四个纳入因子，并记录一个暂不纳入因子。
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr
from statsmodels.stats.outliers_influence import variance_inflation_factor


PROJECT_DIR = Path(__file__).resolve().parents[1]
ROOT_DIR = PROJECT_DIR.parent
SRC_DIR = PROJECT_DIR / "src"
sys.path.insert(0, str(SRC_DIR))

from run_single_factor import DATE_COL, QUANT_COL, STRATEGY_SHEET, SUBJECTIVE_COL  # noqa: E402
from report import save_table  # noqa: E402


TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"

CORE_FACTORS = [
    "相对强弱1000_500_xdqr10_5",
    "时序波动1m_A股指数_周度_000852.SH",
]
AUX_FACTORS = [
    "截面波动_A股指数周度_000852.SH",
    "成交量_A股指数_1w_000852.SH",
]
EXCLUDED_FACTORS = ["成交持仓比_I_FCM_IM"]
MODEL_FACTORS = CORE_FACTORS + AUX_FACTORS
HORIZONS = {
    "下一周收益": "target_1w",
    "未来4周累计相对收益": "target_4w",
}


def rank_ic(x: pd.Series, y: pd.Series) -> tuple[float, float]:
    """Spearman Rank IC 和 p值。"""
    valid = pd.concat([x, y], axis=1).dropna()
    if len(valid) < 3:
        return np.nan, np.nan
    result = spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
    return float(result.correlation), float(result.pvalue)


def ols_hac(y: pd.Series, x: pd.DataFrame, maxlags: int = 4):
    """OLS + HAC标准误。"""
    valid = pd.concat([y, x], axis=1).dropna()
    y_valid = valid.iloc[:, 0]
    x_valid = sm.add_constant(valid.iloc[:, 1:], has_constant="add")
    return sm.OLS(y_valid, x_valid).fit(cov_type="HAC", cov_kwds={"maxlags": maxlags})


def standardize_by_train(train: pd.DataFrame, test: pd.DataFrame, factors: list[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """只使用训练期均值和标准差做标准化，避免测试期信息泄漏。"""
    train_z = train.copy()
    test_z = test.copy()
    for factor in factors:
        mean = train[factor].mean()
        std = train[factor].std()
        if pd.isna(std) or std == 0:
            std = 1.0
        train_z[factor] = (train[factor] - mean) / std
        test_z[factor] = (test[factor] - mean) / std
    return train_z, test_z


def build_targets() -> pd.DataFrame:
    """构造下一周和未来4周累计相对收益。"""
    strategy = pd.read_excel(ROOT_DIR / "data" / "factors.xlsx", sheet_name=STRATEGY_SHEET)
    out = strategy[[DATE_COL, SUBJECTIVE_COL, QUANT_COL]].copy()
    out[DATE_COL] = pd.to_datetime(out[DATE_COL])
    out = out.sort_values(DATE_COL)
    out["current_relative_return"] = out[QUANT_COL] - out[SUBJECTIVE_COL]
    out["target_1w"] = out["current_relative_return"].shift(-1)
    out["target_4w"] = sum(out["current_relative_return"].shift(-i) for i in range(1, 5))
    return out[[DATE_COL, "target_1w", "target_4w"]]


def build_factor_dataset() -> pd.DataFrame:
    """从已有对齐数据中取指定因子，合并两个目标变量。"""
    aligned = pd.read_csv(TABLES_DIR / "full_stock_factor_aligned_data.csv")
    aligned[DATE_COL] = pd.to_datetime(aligned[DATE_COL])
    factors = MODEL_FACTORS + EXCLUDED_FACTORS
    wide = aligned[aligned["factor"].isin(factors)].pivot_table(
        index=DATE_COL,
        columns="factor",
        values="factor_raw",
        aggfunc="first",
    ).reset_index()
    data = pd.merge(wide, build_targets(), on=DATE_COL, how="inner").sort_values(DATE_COL)
    return data


def high_low_spread(df: pd.DataFrame, factor: str, target: str, direction: int) -> float:
    """按训练期方向计算高低分组收益差。"""
    valid = df[[factor, target]].dropna().copy()
    if len(valid) < 20:
        return np.nan
    valid["group"] = pd.qcut(valid[factor], q=5, labels=False, duplicates="drop") + 1
    if valid["group"].nunique() < 2:
        return np.nan
    group_ret = valid.groupby("group")[target].mean()
    low = group_ret.loc[group_ret.index.min()]
    high = group_ret.loc[group_ret.index.max()]
    return float(high - low) if direction >= 0 else float(low - high)


def direction_hit_rate(df: pd.DataFrame, signal: pd.Series, target: str, direction: int = 1) -> float:
    """方向命中率。"""
    valid = pd.concat([signal.rename("signal"), df[target]], axis=1).dropna()
    valid = valid[(valid["signal"] != 0) & (valid[target] != 0)]
    if valid.empty:
        return np.nan
    pred = np.sign(direction * valid["signal"])
    actual = np.sign(valid[target])
    return float((pred == actual).mean())


def single_factor_split_tests(data: pd.DataFrame) -> pd.DataFrame:
    """四个因子的70/30、60/40稳健性检验。"""
    rows = []
    for factor in MODEL_FACTORS:
        for horizon_name, target in HORIZONS.items():
            base = data[[DATE_COL, factor, target]].dropna().sort_values(DATE_COL).reset_index(drop=True)
            for split_name, ratio in [("70/30", 0.7), ("60/40", 0.6)]:
                split_idx = int(len(base) * ratio)
                train = base.iloc[:split_idx].copy()
                test = base.iloc[split_idx:].copy()
                train_z, test_z = standardize_by_train(train, test, [factor])

                train_ic, train_ic_p = rank_ic(train_z[factor], train_z[target])
                direction = 1 if pd.isna(train_ic) or train_ic >= 0 else -1
                test_ic, test_ic_p = rank_ic(test_z[factor], test_z[target])
                train_model = ols_hac(train_z[target], train_z[[factor]])
                test_model = ols_hac(test_z[target], test_z[[factor]]) if len(test_z) >= 10 else None

                for period_name, part, ic, ic_p, model in [
                    ("样本内", train_z, train_ic, train_ic_p, train_model),
                    ("样本外", test_z, test_ic, test_ic_p, test_model),
                ]:
                    rows.append(
                        {
                            "factor": factor,
                            "horizon": horizon_name,
                            "split": split_name,
                            "period": period_name,
                            "sample_count": len(part),
                            "date_start": part[DATE_COL].min().date().isoformat() if len(part) else None,
                            "date_end": part[DATE_COL].max().date().isoformat() if len(part) else None,
                            "train_direction": direction,
                            "rank_ic": ic,
                            "rank_ic_pvalue": ic_p,
                            "ols_beta": model.params.get(factor, np.nan) if model is not None else np.nan,
                            "ols_pvalue": model.pvalues.get(factor, np.nan) if model is not None else np.nan,
                            "high_low_group_spread": high_low_spread(part, factor, target, direction),
                            "direction_hit_rate": direction_hit_rate(part, part[factor], target, direction),
                        }
                    )
    return pd.DataFrame(rows)


def expanding_window_tests(data: pd.DataFrame, min_train_ratio: float = 0.6) -> pd.DataFrame:
    """expanding-window样本外验证。"""
    rows = []
    for factor in MODEL_FACTORS:
        for horizon_name, target in HORIZONS.items():
            base = data[[DATE_COL, factor, target]].dropna().sort_values(DATE_COL).reset_index(drop=True)
            min_train = max(104, int(len(base) * min_train_ratio))
            preds = []
            for idx in range(min_train, len(base)):
                train = base.iloc[:idx].copy()
                test_row = base.iloc[[idx]].copy()
                train_z, test_z = standardize_by_train(train, test_row, [factor])
                model = sm.OLS(train_z[target], sm.add_constant(train_z[[factor]], has_constant="add")).fit()
                pred = float(model.predict(sm.add_constant(test_z[[factor]], has_constant="add")).iloc[0])
                train_ic, _ = rank_ic(train_z[factor], train_z[target])
                direction = 1 if pd.isna(train_ic) or train_ic >= 0 else -1
                preds.append(
                    {
                        DATE_COL: test_row[DATE_COL].iloc[0],
                        "factor": factor,
                        "horizon": horizon_name,
                        "prediction": pred,
                        "actual": float(test_row[target].iloc[0]),
                        "factor_z": float(test_z[factor].iloc[0]),
                        "train_direction": direction,
                    }
                )
            pred_df = pd.DataFrame(preds)
            if pred_df.empty:
                continue
            pred_ic, pred_ic_p = rank_ic(pred_df["prediction"], pred_df["actual"])
            hit = float((np.sign(pred_df["prediction"]) == np.sign(pred_df["actual"])).mean())
            oos_r2 = 1 - ((pred_df["actual"] - pred_df["prediction"]) ** 2).sum() / ((pred_df["actual"] - pred_df["actual"].mean()) ** 2).sum()
            rows.append(
                {
                    "factor": factor,
                    "horizon": horizon_name,
                    "min_train_ratio": min_train_ratio,
                    "oos_count": len(pred_df),
                    "oos_start": pred_df[DATE_COL].min().date().isoformat(),
                    "oos_end": pred_df[DATE_COL].max().date().isoformat(),
                    "prediction_rank_ic": pred_ic,
                    "prediction_rank_ic_pvalue": pred_ic_p,
                    "prediction_direction_hit_rate": hit,
                    "oos_r2": oos_r2,
                }
            )
    return pd.DataFrame(rows)


def multifactor_hac(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """多因子OLS + HAC标准误。"""
    coef_rows = []
    vif_rows = []
    for horizon_name, target in HORIZONS.items():
        base = data[[DATE_COL, target] + MODEL_FACTORS].dropna().sort_values(DATE_COL).reset_index(drop=True)
        train_z, _ = standardize_by_train(base, base.copy(), MODEL_FACTORS)
        model = ols_hac(train_z[target], train_z[MODEL_FACTORS])
        for factor in MODEL_FACTORS:
            coef_rows.append(
                {
                    "horizon": horizon_name,
                    "factor": factor,
                    "beta": model.params.get(factor, np.nan),
                    "t_value_hac": model.tvalues.get(factor, np.nan),
                    "p_value_hac": model.pvalues.get(factor, np.nan),
                    "adjusted_r2": model.rsquared_adj,
                    "sample_count": int(model.nobs),
                }
            )

        x = sm.add_constant(train_z[MODEL_FACTORS], has_constant="add")
        for i, factor in enumerate(x.columns):
            if factor == "const":
                continue
            vif_rows.append(
                {
                    "horizon": horizon_name,
                    "factor": factor,
                    "vif": variance_inflation_factor(x.values, i),
                }
            )
    return pd.DataFrame(coef_rows), pd.DataFrame(vif_rows)


def multifactor_oos(data: pd.DataFrame, split_ratio: float = 0.7) -> pd.DataFrame:
    """多因子样本内/样本外表现。"""
    rows = []
    for horizon_name, target in HORIZONS.items():
        base = data[[DATE_COL, target] + MODEL_FACTORS].dropna().sort_values(DATE_COL).reset_index(drop=True)
        split_idx = int(len(base) * split_ratio)
        train = base.iloc[:split_idx].copy()
        test = base.iloc[split_idx:].copy()
        train_z, test_z = standardize_by_train(train, test, MODEL_FACTORS)
        model = ols_hac(train_z[target], train_z[MODEL_FACTORS])
        for period_name, part in [("样本内", train_z), ("样本外", test_z)]:
            x = sm.add_constant(part[MODEL_FACTORS], has_constant="add")
            pred = model.predict(x)
            pred_ic, pred_ic_p = rank_ic(pred, part[target])
            hit = float((np.sign(pred) == np.sign(part[target])).mean())
            r2 = 1 - ((part[target] - pred) ** 2).sum() / ((part[target] - train_z[target].mean()) ** 2).sum()
            rows.append(
                {
                    "horizon": horizon_name,
                    "split": "70/30",
                    "period": period_name,
                    "sample_count": len(part),
                    "prediction_rank_ic": pred_ic,
                    "prediction_rank_ic_pvalue": pred_ic_p,
                    "direction_hit_rate": hit,
                    "r2_vs_train_mean": r2,
                    "train_adjusted_r2": model.rsquared_adj,
                }
            )
    return pd.DataFrame(rows)


def incremental_r2(data: pd.DataFrame) -> pd.DataFrame:
    """按指定顺序计算单因子加入后的增量解释力。"""
    rows = []
    order = MODEL_FACTORS
    for horizon_name, target in HORIZONS.items():
        used = []
        prev_adj = np.nan
        for factor in order:
            used.append(factor)
            base = data[[target] + used].dropna().copy()
            train_z, _ = standardize_by_train(base, base.copy(), used)
            model = ols_hac(train_z[target], train_z[used])
            adj = model.rsquared_adj
            rows.append(
                {
                    "horizon": horizon_name,
                    "added_factor": factor,
                    "factor_count": len(used),
                    "adjusted_r2": adj,
                    "incremental_adjusted_r2": adj if pd.isna(prev_adj) else adj - prev_adj,
                    "sample_count": int(model.nobs),
                }
            )
            prev_adj = adj
    return pd.DataFrame(rows)


def independent_info(coefs: pd.DataFrame, vif: pd.DataFrame) -> pd.DataFrame:
    """判断多因子模型中仍提供独立信息的因子。"""
    out = coefs.merge(vif, on=["horizon", "factor"], how="left")
    out["独立信息判断"] = np.where(
        (out["p_value_hac"] <= 0.1) & (out["vif"] <= 5),
        "仍提供独立信息",
        np.where(out["vif"] > 5, "受共线性影响较大", "多因子中证据不足"),
    )
    out["说明"] = out.apply(
        lambda row: f"HAC p值={row['p_value_hac']:.4f}，VIF={row['vif']:.2f}，{row['独立信息判断']}。",
        axis=1,
    )
    return out


def factor_layers() -> pd.DataFrame:
    """输出用户指定的因子分层。"""
    rows = []
    for factor in CORE_FACTORS:
        rows.append({"factor": factor, "layer": "核心候选", "说明": "进入稳健性与多因子检验"})
    for factor in AUX_FACTORS:
        rows.append({"factor": factor, "layer": "辅助候选", "说明": "进入稳健性与多因子检验，但解释时作为辅助信息"})
    for factor in EXCLUDED_FACTORS:
        rows.append({"factor": factor, "layer": "暂不纳入", "说明": "本阶段不进入多因子模型"})
    return pd.DataFrame(rows)


def write_report(path: Path, coefs: pd.DataFrame, oos: pd.DataFrame, independent: pd.DataFrame) -> None:
    """中文总结报告。"""
    lines = [
        "# 稳健性与多因子增量检验报告",
        "",
        "## 本阶段范围",
        "",
        "- 不新增或筛选股票因子。",
        "- 不做最终打分、机器学习或PCA。",
        "- 多因子模型仅纳入两个核心候选和两个辅助候选；`成交持仓比_I_FCM_IM` 暂不纳入。",
        "",
        "## 多因子HAC结论",
        "",
    ]
    for _, row in independent.iterrows():
        lines.append(f"- `{row['factor']}`（{row['horizon']}）：{row['说明']}")

    lines.extend(["", "## 多因子样本外表现", ""])
    for _, row in oos[oos["period"] == "样本外"].iterrows():
        lines.append(
            f"- {row['horizon']}：预测Rank IC={row['prediction_rank_ic']:.6f}，"
            f"p值={row['prediction_rank_ic_pvalue']:.6f}，方向命中率={row['direction_hit_rate']:.2%}，"
            f"R2_vs_train_mean={row['r2_vs_train_mean']:.6f}。"
        )
    lines.extend(["", "完整结果见 `robustness_multifactor_report.xlsx`。"])
    path.write_text("\n".join(lines), encoding="utf-8")


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选并设置列宽。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = min(max(max(len("" if cell.value is None else str(cell.value)) for cell in column_cells) + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """导出Excel。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            format_sheet(writer, sheet)


def main() -> None:
    """运行稳健性与多因子增量检验。"""
    data = build_factor_dataset()
    single = single_factor_split_tests(data)
    expanding = expanding_window_tests(data)
    coefs, vif = multifactor_hac(data)
    multi_oos = multifactor_oos(data)
    incr = incremental_r2(data)
    independent = independent_info(coefs, vif)
    layers = factor_layers()

    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    save_table(single, TABLES_DIR / "robustness_single_factor_splits.csv")
    save_table(expanding, TABLES_DIR / "robustness_expanding_oos.csv")
    save_table(coefs, TABLES_DIR / "robustness_multifactor_hac.csv")
    save_table(vif, TABLES_DIR / "robustness_multifactor_vif.csv")
    save_table(multi_oos, TABLES_DIR / "robustness_multifactor_oos.csv")
    save_table(incr, TABLES_DIR / "robustness_incremental_r2.csv")
    save_table(independent, TABLES_DIR / "robustness_independent_info.csv")

    sheets = {
        "项目说明": pd.DataFrame(
            [
                {"项目": "阶段", "说明": "稳健性与多因子增量检验"},
                {"项目": "限制", "说明": "不新增或筛选股票因子；不做最终打分、机器学习或PCA"},
                {"项目": "目标1", "说明": "下一周相对收益"},
                {"项目": "目标2", "说明": "未来4周累计相对收益"},
            ]
        ),
        "因子分层": layers,
        "单因子切分稳健性": single,
        "Expanding样本外": expanding,
        "多因子HAC": coefs,
        "VIF": vif,
        "多因子样本外": multi_oos,
        "增量解释力": incr,
        "独立信息判断": independent,
    }
    export_excel(REPORTS_DIR / "robustness_multifactor_report.xlsx", sheets)
    write_report(REPORTS_DIR / "robustness_multifactor_report.md", coefs, multi_oos, independent)

    print("稳健性与多因子增量检验完成。")
    print(f"Excel：{REPORTS_DIR / 'robustness_multifactor_report.xlsx'}")
    print(f"报告：{REPORTS_DIR / 'robustness_multifactor_report.md'}")


if __name__ == "__main__":
    main()
