"""最终股票因子合成与评分。

只使用用户指定的三个最终保留因子：
- 相对强弱1000_500_xdqr10_5
- 时序波动1m_A股指数_周度_000852.SH
- 截面波动_A股指数周度_000852.SH

暂不纳入：
- 成交量_A股指数_1w_000852.SH
- 成交持仓比_I_FCM_IM

方法：
- expanding-window 真实样本外；
- z-score、方向、权重、分位阈值均只使用历史数据；
- 不做机器学习、PCA或最终打分以外的新增筛选。
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr


PROJECT_DIR = Path(__file__).resolve().parents[1]
ROOT_DIR = PROJECT_DIR.parent
SRC_DIR = PROJECT_DIR / "src"
sys.path.insert(0, str(SRC_DIR))

from report import save_table  # noqa: E402
from run_single_factor import DATE_COL, QUANT_COL, STRATEGY_SHEET, SUBJECTIVE_COL  # noqa: E402


TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"

FINAL_FACTORS = [
    "相对强弱1000_500_xdqr10_5",
    "时序波动1m_A股指数_周度_000852.SH",
    "截面波动_A股指数周度_000852.SH",
]
EXCLUDED_FACTORS = [
    "成交量_A股指数_1w_000852.SH",
    "成交持仓比_I_FCM_IM",
]
METHODS = ["等权合成", "训练期Rank IC加权", "训练期多因子回归系数加权"]
ROLLING_Z_WINDOW = 156
MIN_Z_PERIODS = 52
MIN_TRAIN = 156
WEEKS_PER_YEAR = 52


def rank_ic(x: pd.Series, y: pd.Series) -> tuple[float, float]:
    """Spearman Rank IC 和 p值。"""
    valid = pd.concat([x, y], axis=1).dropna()
    if len(valid) < 3:
        return np.nan, np.nan
    result = spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
    return float(result.correlation), float(result.pvalue)


def build_targets() -> pd.DataFrame:
    """构造下一周和未来4周累计相对收益，并保留下周主观/量化收益。"""
    strategy = pd.read_excel(ROOT_DIR / "data" / "factors.xlsx", sheet_name=STRATEGY_SHEET)
    out = strategy[[DATE_COL, SUBJECTIVE_COL, QUANT_COL]].copy()
    out[DATE_COL] = pd.to_datetime(out[DATE_COL])
    out = out.sort_values(DATE_COL).reset_index(drop=True)
    out["relative_return"] = out[QUANT_COL] - out[SUBJECTIVE_COL]
    out["target_1w"] = out["relative_return"].shift(-1)
    out["target_4w"] = sum(out["relative_return"].shift(-i) for i in range(1, 5))
    out["quant_next"] = out[QUANT_COL].shift(-1)
    out["subjective_next"] = out[SUBJECTIVE_COL].shift(-1)
    out["fifty_fifty_next"] = 0.5 * (out["quant_next"] + out["subjective_next"])
    return out[[DATE_COL, "target_1w", "target_4w", "quant_next", "subjective_next", "fifty_fifty_next"]]


def rolling_zscore(series: pd.Series) -> pd.Series:
    """仅使用历史窗口的滚动z-score；当前值纳入当期历史窗口，不使用未来值。"""
    mean = series.rolling(ROLLING_Z_WINDOW, min_periods=MIN_Z_PERIODS).mean()
    std = series.rolling(ROLLING_Z_WINDOW, min_periods=MIN_Z_PERIODS).std()
    std = std.replace(0, np.nan)
    return (series - mean) / std


def build_dataset() -> pd.DataFrame:
    """读取已有对齐数据并构造滚动z-score。"""
    aligned = pd.read_csv(TABLES_DIR / "full_stock_factor_aligned_data.csv")
    aligned[DATE_COL] = pd.to_datetime(aligned[DATE_COL])
    wide = aligned[aligned["factor"].isin(FINAL_FACTORS)].pivot_table(
        index=DATE_COL,
        columns="factor",
        values="factor_raw",
        aggfunc="first",
    ).reset_index()
    data = pd.merge(wide, build_targets(), on=DATE_COL, how="inner").sort_values(DATE_COL).reset_index(drop=True)
    for factor in FINAL_FACTORS:
        data[f"{factor}__z"] = rolling_zscore(data[factor])
    return data


def train_directions(train: pd.DataFrame) -> dict[str, int]:
    """用训练期Rank IC方向统一为得分越高越利好量化股票。"""
    directions = {}
    for factor in FINAL_FACTORS:
        ic, _ = rank_ic(train[f"{factor}__z"], train["target_1w"])
        directions[factor] = 1 if pd.isna(ic) or ic >= 0 else -1
    return directions


def oriented_frame(df: pd.DataFrame, directions: dict[str, int]) -> pd.DataFrame:
    """按训练期方向调整因子z-score。"""
    out = pd.DataFrame(index=df.index)
    for factor in FINAL_FACTORS:
        out[factor] = directions[factor] * df[f"{factor}__z"]
    return out


def ic_weights(train_x: pd.DataFrame, train_y: pd.Series) -> dict[str, float]:
    """训练期Rank IC绝对值加权。"""
    values = {}
    for factor in FINAL_FACTORS:
        ic, _ = rank_ic(train_x[factor], train_y)
        values[factor] = abs(ic) if pd.notna(ic) else 0.0
    total = sum(values.values())
    if total <= 0:
        return {factor: 1 / len(FINAL_FACTORS) for factor in FINAL_FACTORS}
    return {factor: value / total for factor, value in values.items()}


def regression_weights(train_x: pd.DataFrame, train_y: pd.Series) -> dict[str, float]:
    """训练期多因子回归系数加权。"""
    valid = pd.concat([train_y, train_x], axis=1).dropna()
    if len(valid) < MIN_TRAIN:
        return {factor: 1 / len(FINAL_FACTORS) for factor in FINAL_FACTORS}
    y = valid.iloc[:, 0]
    x = sm.add_constant(valid.iloc[:, 1:], has_constant="add")
    model = sm.OLS(y, x).fit()
    return {factor: float(model.params.get(factor, 0.0)) for factor in FINAL_FACTORS}


def score_by_weights(x: pd.DataFrame, weights: dict[str, float]) -> pd.Series:
    """按权重合成得分。"""
    score = pd.Series(0.0, index=x.index)
    for factor, weight in weights.items():
        score = score + weight * x[factor]
    return score


def classify_regime(score: float, train_scores: pd.Series) -> tuple[str, float, float]:
    """用训练期历史得分分位数划分高/中/低分区。"""
    clean = train_scores.dropna()
    if len(clean) < 50:
        return "中间区：中性", np.nan, np.nan
    low_q = float(clean.quantile(0.3))
    high_q = float(clean.quantile(0.7))
    if score >= high_q:
        return "高分区：量化占优", low_q, high_q
    if score <= low_q:
        return "低分区：主观占优", low_q, high_q
    return "中间区：中性", low_q, high_q


def generate_oos_scores(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """expanding-window生成真实样本外得分、方向、权重和配置收益。"""
    valid_cols = [f"{factor}__z" for factor in FINAL_FACTORS] + ["target_1w", "quant_next", "subjective_next", "fifty_fifty_next"]
    base = data.dropna(subset=valid_cols).sort_values(DATE_COL).reset_index(drop=True)
    rows = []
    weight_rows = []

    for idx in range(MIN_TRAIN, len(base)):
        train = base.iloc[:idx].copy()
        current = base.iloc[[idx]].copy()
        directions = train_directions(train)
        train_x = oriented_frame(train, directions)
        current_x = oriented_frame(current, directions)
        train_y = train["target_1w"]

        weights_by_method = {
            "等权合成": {factor: 1 / len(FINAL_FACTORS) for factor in FINAL_FACTORS},
            "训练期Rank IC加权": ic_weights(train_x, train_y),
            "训练期多因子回归系数加权": regression_weights(train_x, train_y),
        }

        for method, weights in weights_by_method.items():
            train_scores = score_by_weights(train_x, weights)
            score = float(score_by_weights(current_x, weights).iloc[0])
            regime, low_q, high_q = classify_regime(score, train_scores)
            quant_weight = 1.0 if regime.startswith("高分区") else 0.0 if regime.startswith("低分区") else 0.5
            row = current.iloc[0]
            strategy_return = quant_weight * row["quant_next"] + (1 - quant_weight) * row["subjective_next"]
            rows.append(
                {
                    DATE_COL: row[DATE_COL],
                    "method": method,
                    "score": score,
                    "regime": regime,
                    "rolling_p30": low_q,
                    "rolling_p70": high_q,
                    "quant_weight": quant_weight,
                    "target_1w": row["target_1w"],
                    "target_4w": row.get("target_4w", np.nan),
                    "quant_next": row["quant_next"],
                    "subjective_next": row["subjective_next"],
                    "fifty_fifty_next": row["fifty_fifty_next"],
                    "strategy_return": strategy_return,
                }
            )
            for factor in FINAL_FACTORS:
                weight_rows.append(
                    {
                        DATE_COL: row[DATE_COL],
                        "method": method,
                        "factor": factor,
                        "direction": directions[factor],
                        "weight": weights[factor],
                    }
                )
    return pd.DataFrame(rows), pd.DataFrame(weight_rows)


def max_drawdown(returns: pd.Series) -> float:
    """最大回撤。"""
    wealth = (1 + returns.fillna(0)).cumprod()
    drawdown = wealth / wealth.cummax() - 1
    return float(drawdown.min())


def performance_metrics(returns: pd.Series, quant_weight: pd.Series | None = None) -> dict[str, float]:
    """年化收益、波动、夏普、最大回撤和换手率。"""
    ret = returns.dropna()
    if ret.empty:
        return {}
    wealth = float((1 + ret).prod())
    ann_return = wealth ** (WEEKS_PER_YEAR / len(ret)) - 1
    ann_vol = float(ret.std() * np.sqrt(WEEKS_PER_YEAR))
    sharpe = float(ret.mean() / ret.std() * np.sqrt(WEEKS_PER_YEAR)) if ret.std() != 0 else np.nan
    turnover = float(quant_weight.diff().abs().mean()) if quant_weight is not None else 0.0
    return {
        "sample_count": len(ret),
        "cumulative_return": wealth - 1,
        "annualized_return": ann_return,
        "annualized_volatility": ann_vol,
        "sharpe": sharpe,
        "max_drawdown": max_drawdown(ret),
        "turnover": turnover,
    }


def high_low_spread(df: pd.DataFrame, score_col: str, target_col: str) -> float:
    """高低分组收益差。"""
    valid = df[[score_col, target_col]].dropna().copy()
    if len(valid) < 20:
        return np.nan
    valid["group"] = pd.qcut(valid[score_col], q=5, labels=False, duplicates="drop") + 1
    if valid["group"].nunique() < 2:
        return np.nan
    group_ret = valid.groupby("group")[target_col].mean()
    return float(group_ret.loc[group_ret.index.max()] - group_ret.loc[group_ret.index.min()])


def direction_hit_rate(score: pd.Series, target: pd.Series) -> float:
    """方向命中率：sign(得分) 是否等于 sign(下一期相对收益)。"""
    valid = pd.concat([score.rename("score"), target.rename("target")], axis=1).dropna()
    valid = valid[(valid["score"] != 0) & (valid["target"] != 0)]
    if valid.empty:
        return np.nan
    return float((np.sign(valid["score"]) == np.sign(valid["target"])).mean())


def evaluate_methods(scores: pd.DataFrame) -> pd.DataFrame:
    """评价三种合成方法。"""
    rows = []
    for method, part in scores.groupby("method"):
        for target_name in ["target_1w", "target_4w"]:
            ic, ic_p = rank_ic(part["score"], part[target_name])
            rows.append(
                {
                    "method": method,
                    "target": "下一周收益" if target_name == "target_1w" else "未来4周累计相对收益",
                    "rank_ic": ic,
                    "rank_ic_pvalue": ic_p,
                    "direction_hit_rate": direction_hit_rate(part["score"], part[target_name]),
                    "high_low_group_spread": high_low_spread(part, "score", target_name),
                }
            )
    return pd.DataFrame(rows)


def strategy_comparison(scores: pd.DataFrame, best_single: pd.DataFrame) -> pd.DataFrame:
    """比较三种合成、最佳单因子、始终偏量化、始终偏主观和50/50配置。"""
    rows = []
    base = scores[scores["method"] == "等权合成"].sort_values(DATE_COL).reset_index(drop=True)
    benchmarks = {
        "始终偏量化": (base["quant_next"], pd.Series(1.0, index=base.index)),
        "始终偏主观": (base["subjective_next"], pd.Series(0.0, index=base.index)),
        "50/50配置": (base["fifty_fifty_next"], pd.Series(0.5, index=base.index)),
    }
    for method, part in scores.groupby("method"):
        part = part.sort_values(DATE_COL).reset_index(drop=True)
        metrics = performance_metrics(part["strategy_return"], part["quant_weight"])
        metrics["strategy"] = method
        rows.append(metrics)

    best = best_single.sort_values(DATE_COL).reset_index(drop=True)
    metrics = performance_metrics(best["strategy_return"], best["quant_weight"])
    metrics["strategy"] = f"最佳单因子：{best['factor'].iloc[0]}"
    rows.append(metrics)

    for name, (ret, weight) in benchmarks.items():
        metrics = performance_metrics(ret, weight)
        metrics["strategy"] = name
        rows.append(metrics)
    return pd.DataFrame(rows)[
        ["strategy", "sample_count", "cumulative_return", "annualized_return", "annualized_volatility", "sharpe", "max_drawdown", "turnover"]
    ]


def generate_single_factor_scores(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """生成三个单因子的真实样本外得分和配置收益，用于最佳单因子比较。"""
    valid_cols = [f"{factor}__z" for factor in FINAL_FACTORS] + ["target_1w", "quant_next", "subjective_next", "fifty_fifty_next"]
    base = data.dropna(subset=valid_cols).sort_values(DATE_COL).reset_index(drop=True)
    rows = []
    for idx in range(MIN_TRAIN, len(base)):
        train = base.iloc[:idx]
        current = base.iloc[[idx]]
        for factor in FINAL_FACTORS:
            ic, _ = rank_ic(train[f"{factor}__z"], train["target_1w"])
            direction = 1 if pd.isna(ic) or ic >= 0 else -1
            train_score = direction * train[f"{factor}__z"]
            score = float(direction * current[f"{factor}__z"].iloc[0])
            regime, _, _ = classify_regime(score, train_score)
            quant_weight = 1.0 if regime.startswith("高分区") else 0.0 if regime.startswith("低分区") else 0.5
            row = current.iloc[0]
            rows.append(
                {
                    DATE_COL: row[DATE_COL],
                    "factor": factor,
                    "score": score,
                    "target_1w": row["target_1w"],
                    "quant_weight": quant_weight,
                    "strategy_return": quant_weight * row["quant_next"] + (1 - quant_weight) * row["subjective_next"],
                }
            )
    single = pd.DataFrame(rows)
    eval_rows = []
    for factor, part in single.groupby("factor"):
        ic, ic_p = rank_ic(part["score"], part["target_1w"])
        eval_rows.append(
            {
                "factor": factor,
                "rank_ic": ic,
                "rank_ic_pvalue": ic_p,
                "direction_hit_rate": direction_hit_rate(part["score"], part["target_1w"]),
            }
        )
    eval_df = pd.DataFrame(eval_rows).sort_values("rank_ic", ascending=False)
    best_factor = eval_df.iloc[0]["factor"]
    return single[single["factor"] == best_factor].copy(), eval_df


def factor_layer_table() -> pd.DataFrame:
    """最终因子分层。"""
    rows = []
    for factor in FINAL_FACTORS:
        rows.append({"factor": factor, "status": "最终保留", "说明": "进入合成与评分"})
    for factor in EXCLUDED_FACTORS:
        rows.append({"factor": factor, "status": "暂不纳入", "说明": "本阶段不参与合成"})
    return pd.DataFrame(rows)


def write_report(path: Path, method_eval: pd.DataFrame, comparison: pd.DataFrame, single_eval: pd.DataFrame) -> None:
    """中文总结报告。"""
    best_method = comparison.sort_values("sharpe", ascending=False).iloc[0]
    best_single = single_eval.iloc[0]
    lines = [
        "# 股票因子合成与评分报告",
        "",
        "## 本阶段范围",
        "",
        "- 最终保留三个因子：相对强弱、时序波动、截面波动。",
        "- 暂不纳入成交量和成交持仓比。",
        "- 所有z-score、方向、权重和分位数均在expanding-window中只使用历史数据。",
        "- 未使用机器学习、PCA或未来信息。",
        "",
        "## 方向命中率说明",
        "",
        "- 方向命中率按 `sign(样本外得分) == sign(下一期相对收益)` 计算。",
        "- 高分表示更利好量化股票，低分表示更利好主观股票。",
        "- 未来4周累计相对收益存在重叠样本和自相关，因此对应p值和命中率只能作为稳健性参考，不能当作独立周样本解释。",
        "",
        "## 合成方法表现",
        "",
    ]
    for _, row in method_eval[method_eval["target"] == "下一周收益"].iterrows():
        lines.append(
            f"- {row['method']}：样本外Rank IC={row['rank_ic']:.6f}，p值={row['rank_ic_pvalue']:.6f}，"
            f"方向命中率={row['direction_hit_rate']:.2%}，高低分组收益差={row['high_low_group_spread']:.6f}。"
        )
    lines.extend(
        [
            "",
            "## 策略比较",
            "",
            f"- 夏普最高的是 `{best_method['strategy']}`，夏普={best_method['sharpe']:.4f}，累计收益={best_method['cumulative_return']:.2%}。",
            f"- 最佳单因子按样本外Rank IC选择为 `{best_single['factor']}`，Rank IC={best_single['rank_ic']:.6f}。",
            "",
            "完整结果见 `stock_factor_composite_report.xlsx`。",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选、设置列宽。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = min(max(max(len("" if cell.value is None else str(cell.value)) for cell in column_cells) + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """导出Excel。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            format_sheet(writer, sheet)


def main() -> None:
    """运行最终股票因子合成与评分。"""
    data = build_dataset()
    scores, weights = generate_oos_scores(data)
    method_eval = evaluate_methods(scores)
    best_single, single_eval = generate_single_factor_scores(data)
    comparison = strategy_comparison(scores, best_single)
    layers = factor_layer_table()

    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    save_table(scores, TABLES_DIR / "stock_composite_oos_scores.csv")
    save_table(weights, TABLES_DIR / "stock_composite_oos_weights.csv")
    save_table(method_eval, TABLES_DIR / "stock_composite_method_evaluation.csv")
    save_table(single_eval, TABLES_DIR / "stock_composite_single_factor_evaluation.csv")
    save_table(comparison, TABLES_DIR / "stock_composite_strategy_comparison.csv")

    sheets = {
        "项目说明": pd.DataFrame(
            [
                {"项目": "阶段", "说明": "最终股票因子合成与评分"},
                {"项目": "限制", "说明": "不纳入成交量和成交持仓比；不使用机器学习或PCA"},
                {"项目": "样本外方式", "说明": "expanding-window，方向、权重、z-score和分位数仅依赖历史数据"},
                {"项目": "方向命中率", "说明": "sign(样本外得分) == sign(下一期相对收益)"},
                {"项目": "未来4周提示", "说明": "未来4周累计收益存在重叠样本和自相关，仅作稳健性参考"},
            ]
        ),
        "因子分层": layers,
        "样本外得分": scores,
        "样本外权重": weights,
        "方法评价": method_eval,
        "单因子评价": single_eval,
        "策略比较": comparison,
    }
    export_excel(REPORTS_DIR / "stock_factor_composite_report.xlsx", sheets)
    write_report(REPORTS_DIR / "stock_factor_composite_report.md", method_eval, comparison, single_eval)

    print("股票因子合成与评分完成。")
    print(f"Excel：{REPORTS_DIR / 'stock_factor_composite_report.xlsx'}")
    print(f"报告：{REPORTS_DIR / 'stock_factor_composite_report.md'}")


if __name__ == "__main__":
    main()
