"""CTA核心候选复核、去重和样本外验证。

本阶段不做因子合成。
修正滚动稳定性口径：
- rolling_beta_same_sign_ratio：滚动beta与全样本beta同号比例；
- latest_rolling_beta：最近一期滚动beta；
- rolling_beta_sign_flip_count：滚动beta方向翻转次数。
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr


PROJECT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_DIR / "src"
sys.path.insert(0, str(SRC_DIR))

from report import save_table  # noqa: E402


TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"
TARGET_COL = "next_cta_relative_return"
HIGH_CORR_THRESHOLD = 0.7


def rank_ic(x: pd.Series, y: pd.Series) -> tuple[float, float]:
    """Spearman Rank IC 和p值。"""
    valid = pd.concat([x, y], axis=1).dropna()
    if len(valid) < 3:
        return np.nan, np.nan
    result = spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
    return float(result.correlation), float(result.pvalue)


def ols_beta_pvalue(x: pd.Series, y: pd.Series) -> tuple[float, float]:
    """OLS beta和p值。"""
    valid = pd.concat([x, y], axis=1).dropna()
    valid.columns = ["factor", "target"]
    if len(valid) < 5:
        return np.nan, np.nan
    model = sm.OLS(valid["target"], sm.add_constant(valid["factor"], has_constant="add")).fit()
    return float(model.params.get("factor", np.nan)), float(model.pvalues.get("factor", np.nan))


def high_low_spread(df: pd.DataFrame, direction: int) -> float:
    """训练方向下的高低分组收益差。"""
    valid = df[["factor_z_oos", TARGET_COL]].dropna().copy()
    if len(valid) < 20:
        return np.nan
    valid["group"] = pd.qcut(valid["factor_z_oos"], q=5, labels=False, duplicates="drop") + 1
    if valid["group"].nunique() < 2:
        return np.nan
    group_ret = valid.groupby("group")[TARGET_COL].mean()
    low = group_ret.loc[group_ret.index.min()]
    high = group_ret.loc[group_ret.index.max()]
    return float(high - low) if direction >= 0 else float(low - high)


def hit_rate(df: pd.DataFrame, direction: int) -> float:
    """方向命中率。"""
    valid = df[["factor_z_oos", TARGET_COL]].dropna().copy()
    valid = valid[(valid["factor_z_oos"] != 0) & (valid[TARGET_COL] != 0)]
    if valid.empty:
        return np.nan
    pred = np.sign(direction * valid["factor_z_oos"])
    actual = np.sign(valid[TARGET_COL])
    return float((pred == actual).mean())


def factor_type(row: pd.Series) -> tuple[str, str]:
    """分为直接商品CTA、股指期货/权益跨市场代理、待人工确认。"""
    factor = str(row["factor"])
    text = " ".join(str(row.get(col, "")) for col in ["name", "cls2", "cls3", "cls4", "description", "dimension"])
    if any(key in factor for key in ["_IF", "_IC", "_IM", "股指期货", "I_FCM"]) or any(key in text for key in ["股指期货", "期货持仓"]):
        return "股指期货或权益跨市场代理因子", "与股指期货或权益对冲环境相关，可作为跨市场风险偏好/对冲需求代理，不直接反映商品CTA环境。"
    if any(key in factor for key in ["南华", "wind大类商品", "wind商品", "期货因子"]) or any(key in text for key in ["商品", "期货"]):
        return "直接商品CTA因子", "直接来自商品、期货因子或商品指数，较直接反映CTA相关市场环境。"
    return "待人工确认", "列名和目录信息不足，需要人工确认经济含义。"


def rolling_stability(core: pd.DataFrame, rolling: pd.DataFrame) -> pd.DataFrame:
    """计算修正后的滚动稳定性。"""
    rows = []
    for _, row in core.iterrows():
        factor = row["factor"]
        full_beta = row["ols_beta"]
        part = rolling[rolling["factor"] == factor].sort_values("date").copy()
        signs = np.sign(part["beta"].dropna())
        full_sign = np.sign(full_beta)
        if len(signs) == 0 or full_sign == 0 or pd.isna(full_sign):
            same_ratio = np.nan
            flip_count = np.nan
            latest = np.nan
        else:
            same_ratio = float((signs == full_sign).mean())
            flip_count = int((signs != signs.shift(1)).sum() - 1) if len(signs) > 1 else 0
            latest = float(part["beta"].dropna().iloc[-1])
        rows.append(
            {
                "factor": factor,
                "full_sample_beta": full_beta,
                "rolling_beta_same_sign_ratio": same_ratio,
                "latest_rolling_beta": latest,
                "rolling_beta_sign_flip_count": flip_count,
                "rolling_stability_explanation": f"滚动beta与全样本beta同号比例为{same_ratio:.2%}，最近一期beta为{latest:.6f}，方向翻转{flip_count}次。" if pd.notna(same_ratio) else "滚动样本不足，无法判断。",
            }
        )
    return pd.DataFrame(rows)


def economic_group(row: pd.Series) -> str:
    """映射核心候选经济维度。"""
    dim = row["dimension"]
    factor = str(row["factor"])
    if any(key in factor for key in ["_IF", "_IC", "_IM", "股指期货"]):
        return "跨市场代理因子"
    if dim == "趋势强度":
        return "趋势强度"
    if dim == "波动率":
        return "波动率"
    if dim == "成交量":
        return "成交量"
    if dim == "成交持仓比":
        return "成交持仓比"
    if dim in {"基差", "期限结构"}:
        return "商品基差或期限结构"
    return "待人工确认"


def choose_representatives(core: pd.DataFrame) -> pd.DataFrame:
    """按经济维度每组推荐一个代表因子。"""
    rows = []
    for group, part in core.groupby("economic_group"):
        ranked = part.copy()
        ranked["abs_rank_ic"] = ranked["rank_ic"].abs()
        ranked = ranked.sort_values(
            ["rank_ic_pvalue", "ols_p_beta", "rolling_beta_same_sign_ratio", "abs_rank_ic", "sample_count"],
            ascending=[True, True, False, False, False],
        )
        best = ranked.iloc[0].copy()
        evidence_ok = best["rank_ic_pvalue"] <= 0.1 and best["ols_p_beta"] <= 0.1 and best["rolling_beta_same_sign_ratio"] >= 0.6
        best["representative_recommendation"] = "推荐代表" if evidence_ok else "证据不足，暂不强制保留"
        best["representative_reason"] = (
            f"在{group}组内按Rank IC p值、OLS p值、滚动同号比例、|Rank IC|和样本量排序后最优；"
            f"证据状态：{best['representative_recommendation']}。"
        )
        rows.append(best.drop(labels=["abs_rank_ic"]))
    return pd.DataFrame(rows).sort_values("economic_group")


def oos_validate_factor(aligned: pd.DataFrame, factor: str) -> list[dict]:
    """70/30时间顺序样本外验证。"""
    df = aligned[aligned["factor"] == factor].copy()
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").dropna(subset=["factor_raw", TARGET_COL]).reset_index(drop=True)
    split_idx = int(len(df) * 0.7)
    train = df.iloc[:split_idx].copy()
    test = df.iloc[split_idx:].copy()
    mean = train["factor_raw"].mean()
    std = train["factor_raw"].std()
    if pd.isna(std) or std == 0:
        std = 1.0
    train["factor_z_oos"] = (train["factor_raw"] - mean) / std
    test["factor_z_oos"] = (test["factor_raw"] - mean) / std
    train_ic, _ = rank_ic(train["factor_z_oos"], train[TARGET_COL])
    train_direction = 1 if pd.isna(train_ic) or train_ic >= 0 else -1
    rows = []
    period_ics = {}
    for period, part in [("样本内", train), ("样本外", test)]:
        ic, ic_p = rank_ic(part["factor_z_oos"], part[TARGET_COL])
        beta, beta_p = ols_beta_pvalue(part["factor_z_oos"], part[TARGET_COL])
        period_ics[period] = ic
        rows.append(
            {
                "factor": factor,
                "period": period,
                "sample_count": len(part),
                "date_start": part["date"].min().date().isoformat(),
                "date_end": part["date"].max().date().isoformat(),
                "train_direction": train_direction,
                "rank_ic": ic,
                "rank_ic_pvalue": ic_p,
                "ols_beta": beta,
                "ols_pvalue": beta_p,
                "high_low_group_spread": high_low_spread(part, train_direction),
                "direction_hit_rate": hit_rate(part, train_direction),
            }
        )
    same = np.sign(period_ics.get("样本内", np.nan)) == np.sign(period_ics.get("样本外", np.nan))
    for row in rows:
        row["in_out_direction_consistent"] = bool(same)
    return rows


def if_ic_basis_review(core: pd.DataFrame) -> pd.DataFrame:
    """特别核对核心候选中的IF/IC基差类因子。"""
    mask = core["factor"].str.contains("基差", na=False) & core["factor"].str.contains("_IF|_IC", regex=True, na=False)
    out = core.loc[mask, ["factor", "dimension", "rank_ic", "rank_ic_pvalue", "ols_p_beta", "factor_type", "type_reason"]].copy()
    out["why_included"] = "属于期货市场基差/期限结构变量，与CTA研究中的carry、对冲成本和跨市场风险偏好有关。"
    out["direct_commodity_cta_environment"] = "否；IF/IC是股指期货，属于权益跨市场代理，不直接反映商品CTA环境。"
    return out


def write_report(path: Path, core: pd.DataFrame, reps: pd.DataFrame, oos: pd.DataFrame, basis_review: pd.DataFrame) -> None:
    """中文报告。"""
    lines = [
        "# CTA核心候选复核、去重和样本外验证报告",
        "",
        "## 口径修正",
        "",
        "- 不再使用 rolling_beta_positive_ratio 作为统一稳定性指标。",
        "- 新增 rolling_beta_same_sign_ratio：滚动beta与全样本beta同号的比例。",
        "- 同时输出最近一期滚动beta和方向翻转次数。",
        "",
        "## 核心候选分类",
        "",
    ]
    for label, count in core["factor_type"].value_counts().items():
        lines.append(f"- {label}：{count}")
    lines.extend(["", "## IF/IC基差类说明", ""])
    if basis_review.empty:
        lines.append("- 核心候选中没有IF/IC基差类因子。")
    else:
        lines.append("- IF/IC基差类因子被纳入，是因为它们属于期货市场基差/期限结构变量，可代理跨市场风险偏好、对冲成本和权益期货carry。")
        lines.append("- 但它们不是直接商品CTA环境变量，应单独展示，不能与商品基差因子混为一类。")
    lines.extend(["", "## 分类代表因子", ""])
    for _, row in reps.iterrows():
        lines.append(f"- {row['economic_group']}：`{row['factor']}`，{row['representative_reason']}")
    lines.extend(["", "## 样本外验证摘要", ""])
    for _, row in oos[oos["period"] == "样本外"].iterrows():
        lines.append(
            f"- `{row['factor']}`：样本外Rank IC={row['rank_ic']:.6f}，p值={row['rank_ic_pvalue']:.6f}，"
            f"高低组收益差={row['high_low_group_spread']:.6f}，方向命中率={row['direction_hit_rate']:.2%}，"
            f"样本内外方向一致={row['in_out_direction_consistent']}。"
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def format_sheet(writer: pd.ExcelWriter, sheet: str) -> None:
    """格式化工作表。"""
    ws = writer.book[sheet]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, cells in enumerate(ws.columns, start=1):
        width = min(max(max(len("" if c.value is None else str(c.value)) for c in cells) + 2, 10), 70)
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    """运行CTA核心候选复核。"""
    screening = pd.read_csv(TABLES_DIR / "cta_factor_screening.csv")
    rolling = pd.read_csv(TABLES_DIR / "cta_factor_rolling_beta.csv")
    aligned = pd.read_csv(TABLES_DIR / "cta_factor_aligned_data.csv")
    high_corr = pd.read_csv(TABLES_DIR / "cta_factor_high_corr_pairs.csv")

    core = screening[screening["screening_label"] == "核心候选"].copy()
    stability = rolling_stability(core, rolling)
    core = core.merge(stability, on="factor", how="left")
    type_rows = core.apply(factor_type, axis=1)
    core["factor_type"] = [item[0] for item in type_rows]
    core["type_reason"] = [item[1] for item in type_rows]
    core["economic_group"] = core.apply(economic_group, axis=1)
    core["corrected_stability_label"] = np.where(core["rolling_beta_same_sign_ratio"] >= 0.6, "滚动方向较稳定", "滚动方向不稳定")
    core["new_explanation"] = core.apply(
        lambda row: f"{row['factor_type']}；{row['rolling_stability_explanation']}；{row['corrected_stability_label']}。",
        axis=1,
    )
    reps = choose_representatives(core)
    basis_review = if_ic_basis_review(core)
    oos_rows = []
    for factor in core["factor"]:
        oos_rows.extend(oos_validate_factor(aligned, factor))
    oos = pd.DataFrame(oos_rows)

    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    save_table(core, TABLES_DIR / "cta_core_review_regraded.csv")
    save_table(reps, TABLES_DIR / "cta_core_review_representatives.csv")
    save_table(oos, TABLES_DIR / "cta_core_review_oos_validation.csv")
    save_table(basis_review, TABLES_DIR / "cta_core_review_if_ic_basis.csv")

    xlsx = REPORTS_DIR / "cta_core_candidate_review_oos_report.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        sheets = {
            "项目说明": pd.DataFrame(
                [
                    {"项目": "阶段", "说明": "CTA核心候选复核、去重和样本外验证"},
                    {"项目": "滚动稳定性修正", "说明": "使用rolling_beta_same_sign_ratio，不再用rolling_beta_positive_ratio统一判断"},
                    {"项目": "限制", "说明": "不做因子合成"},
                ]
            ),
            "核心候选复核": core,
            "IF_IC基差核对": basis_review,
            "分类代表因子": reps,
            "样本外验证": oos,
            "高相关因子对": high_corr,
        }
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            format_sheet(writer, sheet)
    write_report(REPORTS_DIR / "cta_core_candidate_review_oos_report.md", core, reps, oos, basis_review)

    print("CTA核心候选复核、去重和样本外验证完成。")
    print(f"核心候选数：{len(core)}")
    print(f"代表因子数：{len(reps)}")
    print(f"Excel：{xlsx}")


if __name__ == "__main__":
    main()
