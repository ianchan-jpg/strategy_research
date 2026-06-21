"""导出因子筛选 Excel 报告。

把 batch_factor_summary.csv、batch_factor_screening.csv 和
batch_factor_high_corr_pairs.csv 合并到一个 Excel 工作簿中。
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"


def explain_factor(row: pd.Series) -> str:
    """用一句中文解释单个因子的初筛结果，方便新手阅读。"""
    label = row["screening_label"]
    rank_ic = row["rank_ic"]
    rank_ic_p = row["rank_ic_pvalue"]
    beta = row["ols_beta"]
    beta_p = row["ols_p_beta"]
    rolling_ratio = row["rolling_beta_positive_ratio"]
    monotonic = row["group_return_monotonic"]

    direction = "正向" if rank_ic >= 0 else "反向"
    monotonic_text = "分组较单调" if bool(monotonic) else "分组不单调"

    if label == "保留":
        return (
            f"该因子与下一期相对收益呈{direction}关系，Rank IC p值为{rank_ic_p:.3f}，"
            f"回归beta为{beta:.4f}且p值为{beta_p:.3f}，滚动beta为正比例{rolling_ratio:.1%}，可进入后续复核。"
        )
    if label == "待复核":
        return (
            f"该因子有一定{direction}信号，但统计证据还不完整，Rank IC p值为{rank_ic_p:.3f}，"
            f"{monotonic_text}，建议结合经济含义继续检查。"
        )
    if label == "重复":
        return (
            f"该因子自身结果一般或与其他因子高度相似，Rank IC为{rank_ic:.4f}，"
            "暂不删除，但应优先和高相关因子一起比较。"
        )
    return (
        f"该因子当前证据不足，Rank IC为{rank_ic:.4f}且p值为{rank_ic_p:.3f}，"
        f"回归显著性也偏弱，暂不作为核心候选。"
    )


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """设置冻结首行、筛选、列宽和表头样式。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        width = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    """生成 factor_screening_report.xlsx。"""
    summary = pd.read_csv(TABLES_DIR / "batch_factor_summary.csv")
    screening = pd.read_csv(TABLES_DIR / "batch_factor_screening.csv")
    high_corr = pd.read_csv(TABLES_DIR / "batch_factor_high_corr_pairs.csv")

    label_order = {"保留": 1, "待复核": 2, "重复": 3, "证据不足": 4}
    report = screening.copy()
    report["rank_ic_abs"] = report["rank_ic"].abs()
    report["label_order"] = report["screening_label"].map(label_order).fillna(99)
    report["新手解释"] = report.apply(explain_factor, axis=1)
    report = report.sort_values(["label_order", "rank_ic_abs"], ascending=[True, False]).drop(columns=["label_order"])

    output_path = REPORTS_DIR / "factor_screening_report.xlsx"
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="因子筛选报告", index=False)
        summary.to_excel(writer, sheet_name="批量测试汇总", index=False)
        high_corr.to_excel(writer, sheet_name="高相关因子对", index=False)

        for sheet_name in ["因子筛选报告", "批量测试汇总", "高相关因子对"]:
            format_sheet(writer, sheet_name)

    print(f"已生成：{output_path}")


if __name__ == "__main__":
    main()
