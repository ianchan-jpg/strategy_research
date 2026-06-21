"""研究维度覆盖审计。

只基于现有股票组全量检验结果做覆盖审计和差距分析。
不新增模型，不重新筛选。
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"

DIMENSIONS = [
    "市场活跃度",
    "大小盘风格",
    "市场宽度",
    "截面分化",
    "时序波动",
    "行业轮动",
    "龙头集中度",
    "赚钱效应",
    "因子有效性",
    "情绪或拥挤度",
]

FINAL_FACTORS = {
    "相对强弱1000_500_xdqr10_5": "大小盘风格",
    "时序波动1m_A股指数_周度_000852.SH": "时序波动",
    "截面波动_A股指数周度_000852.SH": "截面分化",
}


def text_blob(row: pd.Series) -> str:
    """合并字段用于关键词匹配。"""
    cols = ["factor", "name", "cls3", "cls4", "description", "formula", "code"]
    return " ".join("" if col not in row or pd.isna(row[col]) else str(row[col]) for col in cols)


def infer_dimension(row: pd.Series) -> str | None:
    """把因子映射到原始研究维度。"""
    text = text_blob(row)
    factor = str(row.get("factor", ""))
    name = str(row.get("name", ""))

    if any(key in text for key in ["成交量_A股指数", "成交额_A股指数", "市场成交量", "市场成交额"]):
        return "市场活跃度"
    if "相对强弱" in text:
        return "大小盘风格"
    if "赚钱效应" in text:
        return "赚钱效应"
    if "时序波动" in text:
        return "时序波动"
    if "截面波动变化" in text:
        return "市场宽度"
    if "截面波动" in text:
        return "截面分化"
    if "industry_style_cross" in factor or "行业" in name or "中信风格" in text:
        return "行业轮动"
    if "instable" in factor or "不稳定性" in text or "barra" in text:
        return "因子有效性"
    if any(key in text for key in ["VIX", "拥挤度", "成交持仓比", "持仓量", "基差率", "期货持仓", "对冲成本"]):
        return "情绪或拥挤度"
    if any(key in text for key in ["龙头", "集中度", "强势股"]):
        return "龙头集中度"
    return None


def main_reason(rows: pd.DataFrame) -> str:
    """总结未保留主要原因。"""
    if rows.empty:
        return "未找到合适指标"
    if "factor" in rows.columns and rows["factor"].isin(FINAL_FACTORS).any():
        return "已有最终保留因子覆盖"
    if "exclude_reason" in rows.columns and rows["exclude_reason"].notna().any():
        top = rows["exclude_reason"].dropna().value_counts()
        if not top.empty:
            return top.index[0]
    if "screening_label" in rows.columns:
        labels = rows["screening_label"].fillna("").value_counts()
        if labels.get("核心候选", 0) == 0 and labels.get("弱候选", 0) == 0 and labels.get("组内代表", 0) == 0:
            return "已测试但统计证据不足"
        if labels.get("核心候选", 0) > 0 and not rows["factor"].isin(FINAL_FACTORS).any():
            return "有核心候选但未进入最终三因子，主要因去重或稳健性比较未优先保留"
    return "待人工复核"


def coverage_status(original_count: int, tested_count: int, excluded_count: int, core_count: int, final_count: int) -> str:
    """判断覆盖状态。"""
    if original_count == 0:
        return "未找到合适指标"
    if tested_count == 0 and excluded_count > 0:
        return "数据不足"
    if tested_count == 0:
        return "尚未测试"
    if final_count > 0:
        return "已覆盖"
    if core_count > 0:
        return "已充分测试但未最终保留"
    return "已充分测试但未通过"


def build_audit() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """构造覆盖审计结果。"""
    candidates = pd.read_csv(TABLES_DIR / "full_stock_factor_candidates.csv")
    exclusions = pd.read_csv(TABLES_DIR / "full_stock_factor_exclusion_list.csv")
    screening = pd.read_csv(TABLES_DIR / "full_stock_factor_screening.csv")

    candidates["dimension"] = candidates.apply(infer_dimension, axis=1)
    exclusions["dimension"] = exclusions.apply(infer_dimension, axis=1)
    screening["dimension"] = screening.apply(infer_dimension, axis=1)
    screening["is_final_factor"] = screening["factor"].isin(FINAL_FACTORS)

    detail = pd.concat(
        [
            screening.assign(source="已测试"),
            exclusions.assign(source="已排除"),
        ],
        ignore_index=True,
        sort=False,
    )

    rows = []
    for dimension in DIMENSIONS:
        tested = screening[screening["dimension"] == dimension]
        excluded = exclusions[exclusions["dimension"] == dimension]
        original = pd.concat([tested[["factor"]], excluded[["factor"]]], ignore_index=True).drop_duplicates()
        core = tested[tested["screening_label"] == "核心候选"]
        final = tested[tested["factor"].isin(FINAL_FACTORS)]
        rows.append(
            {
                "研究维度": dimension,
                "原始候选因子数": len(original),
                "实际测试因子数": len(tested),
                "被排除因子数": len(excluded),
                "核心候选数": len(core),
                "最终保留数": len(final),
                "覆盖状态": coverage_status(len(original), len(tested), len(excluded), len(core), len(final)),
                "未保留的主要原因": main_reason(pd.concat([tested, excluded], ignore_index=True, sort=False)),
                "最终保留因子": " | ".join(final["factor"].tolist()),
                "核心候选示例": " | ".join(core["factor"].head(5).tolist()),
            }
        )
    audit = pd.DataFrame(rows)

    final_coverage = pd.DataFrame(
        [
            {"最终因子": factor, "覆盖维度": dimension, "说明": "进入最终股票因子合成与评分"}
            for factor, dimension in FINAL_FACTORS.items()
        ]
    )
    return audit, detail, final_coverage


def write_report(path: Path, audit: pd.DataFrame, final_coverage: pd.DataFrame) -> None:
    """中文覆盖审计报告。"""
    covered = audit[audit["最终保留数"] > 0]["研究维度"].tolist()
    gap = audit[audit["最终保留数"] == 0]["研究维度"].tolist()
    no_indicator = audit[audit["覆盖状态"] == "未找到合适指标"]["研究维度"].tolist()
    failed = audit[audit["覆盖状态"] == "已充分测试但未通过"]["研究维度"].tolist()
    not_final = audit[audit["覆盖状态"] == "已充分测试但未最终保留"]["研究维度"].tolist()

    lines = [
        "# 研究维度覆盖审计报告",
        "",
        "## 结论摘要",
        "",
        f"- 当前最终3个因子覆盖维度：{'、'.join(covered) if covered else '无'}。",
        f"- 尚未由最终因子覆盖的维度：{'、'.join(gap) if gap else '无'}。",
        "- 当前3个最终因子不能覆盖全部原始研究维度，主要覆盖大小盘风格、时序波动和截面分化。",
        "",
        "## 差距分析",
        "",
        f"- 已充分测试但未通过：{'、'.join(failed) if failed else '无'}。",
        f"- 已测试且有候选但未最终保留：{'、'.join(not_final) if not_final else '无'}。",
        f"- 未找到合适指标：{'、'.join(no_indicator) if no_indicator else '无'}。",
        "",
        "## 建议",
        "",
        "- 短期不建议为了覆盖维度而强行加入未通过检验的因子。",
        "- 市场活跃度和情绪/拥挤度已有较多测试和核心候选，但最终因子合成阶段因稳健性或独立信息考虑未保留。",
        "- 行业轮动、赚钱效应、因子有效性等维度已有测试但证据不足或未最终保留，后续可在新数据或更长样本中复核。",
        "- 龙头集中度当前在周度权益量化候选中缺少明确匹配指标，需要补充数据定义后再测试。",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选并设置列宽。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = min(max(max(len("" if cell.value is None else str(cell.value)) for cell in column_cells) + 2, 10), 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """导出Excel。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            format_sheet(writer, sheet)


def main() -> None:
    """运行研究维度覆盖审计。"""
    audit, detail, final_coverage = build_audit()
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    save_path = TABLES_DIR / "research_dimension_coverage_audit.csv"
    audit.to_csv(save_path, index=False, encoding="utf-8-sig")
    detail.to_csv(TABLES_DIR / "research_dimension_factor_detail.csv", index=False, encoding="utf-8-sig")
    final_coverage.to_csv(TABLES_DIR / "research_dimension_final_factor_coverage.csv", index=False, encoding="utf-8-sig")

    export_excel(
        REPORTS_DIR / "research_dimension_coverage_audit.xlsx",
        {
            "覆盖审计": audit,
            "因子明细": detail,
            "最终因子覆盖": final_coverage,
        },
    )
    write_report(REPORTS_DIR / "research_dimension_coverage_audit.md", audit, final_coverage)

    print("研究维度覆盖审计完成。")
    print(f"Excel：{REPORTS_DIR / 'research_dimension_coverage_audit.xlsx'}")
    print(f"报告：{REPORTS_DIR / 'research_dimension_coverage_audit.md'}")


if __name__ == "__main__":
    main()
