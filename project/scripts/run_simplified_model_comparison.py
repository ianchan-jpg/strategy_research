"""精简组合比较。

不新增因子，不使用机器学习、PCA。
比较五种方案：
1. 仅截面波动
2. 截面波动 + 时序波动
3. 截面波动 + 大小盘相对强弱
4. 三因子等权
5. 三因子Rank IC加权
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr


PROJECT_DIR = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_DIR / "src"
sys.path.insert(0, str(SRC_DIR))

from report import save_table  # noqa: E402
from run_stock_factor_composite import (  # noqa: E402
    DATE_COL,
    FINAL_FACTORS,
    MIN_TRAIN,
    WEEKS_PER_YEAR,
    build_dataset,
    classify_regime,
    direction_hit_rate,
    high_low_spread,
    max_drawdown,
    rank_ic,
    train_directions,
    oriented_frame,
    ic_weights,
)


TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"

FACTOR_SIZE = "相对强弱1000_500_xdqr10_5"
FACTOR_TS_VOL = "时序波动1m_A股指数_周度_000852.SH"
FACTOR_CS_VOL = "截面波动_A股指数周度_000852.SH"

SCHEMES = {
    "仅截面波动": {"factors": [FACTOR_CS_VOL], "weight_method": "equal"},
    "截面波动+时序波动": {"factors": [FACTOR_CS_VOL, FACTOR_TS_VOL], "weight_method": "equal"},
    "截面波动+大小盘相对强弱": {"factors": [FACTOR_CS_VOL, FACTOR_SIZE], "weight_method": "equal"},
    "三因子等权": {"factors": [FACTOR_SIZE, FACTOR_TS_VOL, FACTOR_CS_VOL], "weight_method": "equal"},
    "三因子Rank IC加权": {"factors": [FACTOR_SIZE, FACTOR_TS_VOL, FACTOR_CS_VOL], "weight_method": "rank_ic"},
}
COST_BPS_LIST = [0, 5, 10, 20]
HOLDING_PERIODS = [1, 2, 4]


def score_by_weights(x: pd.DataFrame, weights: dict[str, float]) -> pd.Series:
    """按权重合成得分。"""
    score = pd.Series(0.0, index=x.index)
    for factor, weight in weights.items():
        score = score + weight * x[factor]
    return score


def scheme_weights(train_x: pd.DataFrame, train_y: pd.Series, factors: list[str], method: str) -> dict[str, float]:
    """根据方案确定权重。"""
    if method == "rank_ic":
        values = {}
        for factor in factors:
            ic, _ = rank_ic(train_x[factor], train_y)
            values[factor] = abs(ic) if pd.notna(ic) else 0.0
        total = sum(values.values())
        if total > 0:
            return {factor: values[factor] / total for factor in factors}
    return {factor: 1 / len(factors) for factor in factors}


def generate_scheme_scores(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """使用expanding-window生成五种方案的真实样本外得分。"""
    required = [f"{factor}__z" for factor in FINAL_FACTORS] + ["target_1w", "target_4w", "quant_next", "subjective_next", "fifty_fifty_next"]
    base = data.dropna(subset=required).sort_values(DATE_COL).reset_index(drop=True)
    rows = []
    weight_rows = []
    for idx in range(MIN_TRAIN, len(base)):
        train = base.iloc[:idx].copy()
        current = base.iloc[[idx]].copy()
        directions = train_directions(train)
        all_train_x = oriented_frame(train, directions)
        all_current_x = oriented_frame(current, directions)
        train_y = train["target_1w"]

        for scheme, cfg in SCHEMES.items():
            factors = cfg["factors"]
            train_x = all_train_x[factors]
            current_x = all_current_x[factors]
            weights = scheme_weights(train_x, train_y, factors, cfg["weight_method"])
            train_scores = score_by_weights(train_x, weights)
            score = float(score_by_weights(current_x, weights).iloc[0])
            regime, low_q, high_q = classify_regime(score, train_scores)
            quant_weight_raw = 1.0 if regime.startswith("高分区") else 0.0 if regime.startswith("低分区") else 0.5
            row = current.iloc[0]
            rows.append(
                {
                    DATE_COL: row[DATE_COL],
                    "scheme": scheme,
                    "score": score,
                    "regime": regime,
                    "rolling_p30": low_q,
                    "rolling_p70": high_q,
                    "quant_weight_raw": quant_weight_raw,
                    "target_1w": row["target_1w"],
                    "target_4w": row["target_4w"],
                    "quant_next": row["quant_next"],
                    "subjective_next": row["subjective_next"],
                    "fifty_fifty_next": row["fifty_fifty_next"],
                }
            )
            for factor in factors:
                weight_rows.append(
                    {
                        DATE_COL: row[DATE_COL],
                        "scheme": scheme,
                        "factor": factor,
                        "direction": directions[factor],
                        "weight": weights[factor],
                    }
                )
    return pd.DataFrame(rows), pd.DataFrame(weight_rows)


def apply_holding_period(scores: pd.DataFrame, holding_period: int) -> pd.DataFrame:
    """把原始目标权重转换为最短持有期约束下的实际权重。"""
    rows = []
    for scheme, part in scores.groupby("scheme"):
        part = part.sort_values(DATE_COL).reset_index(drop=True).copy()
        actual_weights = []
        current_weight = None
        weeks_held = 0
        for raw_weight in part["quant_weight_raw"]:
            if current_weight is None:
                current_weight = raw_weight
                weeks_held = 1
            elif weeks_held >= holding_period and raw_weight != current_weight:
                current_weight = raw_weight
                weeks_held = 1
            else:
                weeks_held += 1
            actual_weights.append(current_weight)
        part["holding_period"] = holding_period
        part["quant_weight"] = actual_weights
        part["turnover"] = part["quant_weight"].diff().abs().fillna(0.0)
        part["gross_return"] = part["quant_weight"] * part["quant_next"] + (1 - part["quant_weight"]) * part["subjective_next"]
        rows.append(part)
    return pd.concat(rows, ignore_index=True)


def apply_costs(held_scores: pd.DataFrame) -> pd.DataFrame:
    """加入三档交易成本，单位bps。"""
    rows = []
    for cost_bps in COST_BPS_LIST:
        part = held_scores.copy()
        part["cost_bps"] = cost_bps
        part["cost"] = part["turnover"] * cost_bps / 10000
        part["net_return"] = part["gross_return"] - part["cost"]
        rows.append(part)
    return pd.concat(rows, ignore_index=True)


def annualized_metrics(ret: pd.Series, turnover: pd.Series) -> dict[str, float]:
    """绩效指标。"""
    clean = ret.dropna()
    if clean.empty:
        return {}
    wealth = float((1 + clean).prod())
    ann_return = wealth ** (WEEKS_PER_YEAR / len(clean)) - 1
    ann_vol = float(clean.std() * np.sqrt(WEEKS_PER_YEAR))
    sharpe = float(clean.mean() / clean.std() * np.sqrt(WEEKS_PER_YEAR)) if clean.std() != 0 else np.nan
    return {
        "sample_count": len(clean),
        "cumulative_return": wealth - 1,
        "annualized_return": ann_return,
        "annualized_volatility": ann_vol,
        "sharpe": sharpe,
        "max_drawdown": max_drawdown(clean),
        "turnover": float(turnover.mean()),
    }


def evaluate_scheme_costs(cost_scores: pd.DataFrame) -> pd.DataFrame:
    """比较方案、成本和持有期。"""
    rows = []
    for (scheme, holding_period, cost_bps), part in cost_scores.groupby(["scheme", "holding_period", "cost_bps"]):
        ic, ic_p = rank_ic(part["score"], part["target_1w"])
        metrics = annualized_metrics(part["net_return"], part["turnover"])
        rows.append(
            {
                "scheme": scheme,
                "holding_period": holding_period,
                "cost_bps": cost_bps,
                "rank_ic": ic,
                "rank_ic_pvalue": ic_p,
                "direction_hit_rate": direction_hit_rate(part["score"], part["target_1w"]),
                "high_low_group_spread": high_low_spread(part, "score", "target_1w"),
                **metrics,
            }
        )
    return pd.DataFrame(rows).sort_values(["cost_bps", "holding_period", "sharpe"], ascending=[True, True, False])


def market_stage(row: pd.Series) -> str:
    """按量化与主观平均收益划分市场阶段。"""
    avg = 0.5 * (row["quant_next"] + row["subjective_next"])
    rel = row["target_1w"]
    if avg >= 0 and rel >= 0:
        return "上涨且量化占优"
    if avg >= 0 and rel < 0:
        return "上涨但主观占优"
    if avg < 0 and rel >= 0:
        return "下跌但量化占优"
    return "下跌且主观占优"


def stage_stability(cost_scores: pd.DataFrame) -> pd.DataFrame:
    """不同市场阶段稳定性。默认用0bps、1周持有期观察信号本身。"""
    base = cost_scores[(cost_scores["cost_bps"] == 0) & (cost_scores["holding_period"] == 1)].copy()
    base["market_stage"] = base.apply(market_stage, axis=1)
    rows = []
    for (scheme, stage), part in base.groupby(["scheme", "market_stage"]):
        ic, ic_p = rank_ic(part["score"], part["target_1w"])
        metrics = annualized_metrics(part["net_return"], part["turnover"])
        rows.append(
            {
                "scheme": scheme,
                "market_stage": stage,
                "rank_ic": ic,
                "rank_ic_pvalue": ic_p,
                "direction_hit_rate": direction_hit_rate(part["score"], part["target_1w"]),
                **metrics,
            }
        )
    return pd.DataFrame(rows).sort_values(["scheme", "market_stage"])


def recommendations(evaluation: pd.DataFrame) -> pd.DataFrame:
    """推荐主方案、备选方案、基准方案。用10bps、2周持有期作为实用口径。"""
    practical = evaluation[(evaluation["cost_bps"] == 10) & (evaluation["holding_period"] == 2)].copy()
    practical = practical.sort_values(["sharpe", "rank_ic", "max_drawdown"], ascending=[False, False, False])
    benchmark_scheme = "仅截面波动"
    main = practical.iloc[0]
    backup_pool = practical[~practical["scheme"].isin([main["scheme"], benchmark_scheme])]
    backup = backup_pool.iloc[0] if not backup_pool.empty else practical[practical["scheme"] != main["scheme"]].iloc[0]
    benchmark = practical[practical["scheme"] == benchmark_scheme].iloc[0]
    return pd.DataFrame(
        [
            {"role": "主方案", "scheme": main["scheme"], "reason": "在10bps成本、2周持有期口径下夏普最高，且Rank IC为正。"},
            {"role": "备选方案", "scheme": backup["scheme"], "reason": "排除主方案和单因子基准后，在同一实用口径下表现相对较好。"},
            {"role": "基准方案", "scheme": benchmark["scheme"], "reason": "仅使用此前最佳单因子截面波动，作为最简基准方案。"},
        ]
    )


def write_report(path: Path, evaluation: pd.DataFrame, rec: pd.DataFrame, stage: pd.DataFrame) -> None:
    """中文总结报告。"""
    practical = evaluation[(evaluation["cost_bps"] == 10) & (evaluation["holding_period"] == 2)].sort_values("sharpe", ascending=False)
    lines = [
        "# 精简组合比较报告",
        "",
        "## 本阶段范围",
        "",
        "- 不新增因子，不使用机器学习、PCA。",
        "- 使用相同expanding-window样本外框架。",
        "- 比较5种方案、3档交易成本和3种最短持有期。",
        "",
        "## 推荐方案",
        "",
    ]
    for _, row in rec.iterrows():
        lines.append(f"- {row['role']}：`{row['scheme']}`。{row['reason']}")
    lines.extend(["", "## 10bps成本、2周持有期表现", ""])
    for _, row in practical.iterrows():
        lines.append(
            f"- `{row['scheme']}`：Rank IC={row['rank_ic']:.6f}，年化收益={row['annualized_return']:.2%}，"
            f"夏普={row['sharpe']:.4f}，最大回撤={row['max_drawdown']:.2%}，换手率={row['turnover']:.4f}。"
        )
    lines.extend(["", "## 市场阶段稳定性", ""])
    best_scheme = rec.loc[rec["role"] == "主方案", "scheme"].iloc[0]
    for _, row in stage[stage["scheme"] == best_scheme].iterrows():
        lines.append(
            f"- `{best_scheme}` 在 `{row['market_stage']}`：Rank IC={row['rank_ic']:.6f}，方向命中率={row['direction_hit_rate']:.2%}，夏普={row['sharpe']:.4f}。"
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选并设置列宽。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = min(max(max(len("" if cell.value is None else str(cell.value)) for cell in column_cells) + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """导出Excel。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            format_sheet(writer, sheet)


def main() -> None:
    """运行精简组合比较。"""
    data = build_dataset()
    scores, weights = generate_scheme_scores(data)
    held = pd.concat([apply_holding_period(scores, hp) for hp in HOLDING_PERIODS], ignore_index=True)
    cost_scores = apply_costs(held)
    evaluation = evaluate_scheme_costs(cost_scores)
    stage = stage_stability(cost_scores)
    rec = recommendations(evaluation)

    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    save_table(scores, TABLES_DIR / "simplified_model_oos_scores.csv")
    save_table(weights, TABLES_DIR / "simplified_model_oos_weights.csv")
    save_table(cost_scores, TABLES_DIR / "simplified_model_cost_holding_returns.csv")
    save_table(evaluation, TABLES_DIR / "simplified_model_evaluation.csv")
    save_table(stage, TABLES_DIR / "simplified_model_market_stage.csv")
    save_table(rec, TABLES_DIR / "simplified_model_recommendations.csv")

    sheets = {
        "项目说明": pd.DataFrame(
            [
                {"项目": "阶段", "说明": "精简组合比较"},
                {"项目": "限制", "说明": "不新增因子，不使用机器学习或PCA"},
                {"项目": "方案数量", "说明": "5种"},
                {"项目": "交易成本", "说明": "0bps、5bps、10bps、20bps"},
                {"项目": "最短持有期", "说明": "1周、2周、4周"},
            ]
        ),
        "方案定义": pd.DataFrame(
            [{"scheme": k, "factors": " | ".join(v["factors"]), "weight_method": v["weight_method"]} for k, v in SCHEMES.items()]
        ),
        "样本外得分": scores,
        "权重方向": weights,
        "成本持有期收益": cost_scores,
        "方案评价": evaluation,
        "市场阶段稳定性": stage,
        "最终推荐": rec,
    }
    export_excel(REPORTS_DIR / "simplified_model_comparison.xlsx", sheets)
    write_report(REPORTS_DIR / "simplified_model_comparison.md", evaluation, rec, stage)

    print("精简组合比较完成。")
    print(f"Excel：{REPORTS_DIR / 'simplified_model_comparison.xlsx'}")
    print(f"报告：{REPORTS_DIR / 'simplified_model_comparison.md'}")


if __name__ == "__main__":
    main()
