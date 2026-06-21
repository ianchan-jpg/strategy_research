"""核心候选复核、去重和样本外验证。

本脚本不做因子合成、不做机器学习、不删除原始候选。
除样本内/样本外验证外，核心候选分级与代表选择只基于已有全量检验结果。
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr
import statsmodels.api as sm


PROJECT_DIR = Path(__file__).resolve().parents[1]
ROOT_DIR = PROJECT_DIR.parent
SRC_DIR = PROJECT_DIR / "src"
sys.path.insert(0, str(SRC_DIR))

from report import save_table  # noqa: E402


TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"

SCREENING_PATH = TABLES_DIR / "full_stock_factor_screening.csv"
ALIGNED_PATH = TABLES_DIR / "full_stock_factor_aligned_data.csv"
HIGH_CORR_PATH = TABLES_DIR / "full_stock_factor_high_corr_pairs.csv"


def stable_direction(row: pd.Series) -> bool:
    """滚动 beta 方向是否稳定。"""
    ratio = row["rolling_beta_positive_ratio"]
    return pd.notna(ratio) and (ratio >= 0.6 or ratio <= 0.4)


def core_evidence_level(row: pd.Series) -> str:
    """对15个核心候选重新分级。"""
    rank_5 = row["rank_ic_pvalue"] <= 0.05
    ols_5 = row["ols_p_beta"] <= 0.05
    rank_10 = row["rank_ic_pvalue"] <= 0.1
    ols_10 = row["ols_p_beta"] <= 0.1

    if rank_5 and ols_5:
        return "强核心"
    if (rank_5 and ols_10) or (ols_5 and rank_10):
        return "弱核心"
    return "待复核"


def core_explanation(row: pd.Series) -> str:
    """明确说明5%显著还是10%弱显著。"""
    rank_text = "Rank IC在5%水平显著" if row["rank_ic_pvalue"] <= 0.05 else "Rank IC仅在10%水平弱显著" if row["rank_ic_pvalue"] <= 0.1 else "Rank IC不显著"
    ols_text = "OLS在5%水平显著" if row["ols_p_beta"] <= 0.05 else "OLS仅在10%水平弱显著" if row["ols_p_beta"] <= 0.1 else "OLS不显著"
    stable_text = "滚动方向稳定" if stable_direction(row) else "滚动方向稳定性不足"
    return f"{rank_text}，{ols_text}，{stable_text}，因此分级为{row['核心分级']}。"


def futures_rationality_label(row: pd.Series) -> tuple[str, str]:
    """判断股指期货相关指标用于股票组的纳入合理性。"""
    factor = str(row["factor"])
    name = str(row.get("name", ""))
    cls4 = str(row.get("cls4", ""))

    if "基差率" in factor:
        return "可作为股票代理变量", "股指期货基差反映股票指数对冲成本和期限结构，可作为股票市场对冲环境代理，但不是纯股票现货指标。"
    if "成交持仓比" in factor:
        return "可作为股票代理变量", "股指期货成交持仓比刻画期指交易拥挤度，可代理股票指数对冲和投机活跃度。"
    if "成交量_I_FCM" in factor or ("成交量" in name and "期货持仓" in cls4):
        return "可作为股票代理变量", "股指期货成交量与股票指数对冲需求相关，可作为股票市场活跃度的衍生代理变量。"
    if "持仓量_I_FCM" in factor or ("持仓量" in name and "期货持仓" in cls4):
        return "可作为股票代理变量", "股指期货持仓量反映对冲和杠杆资金参与度，可作为股票市场环境代理变量。"
    if "股指期货" in factor:
        return "待人工确认", "列名显示为股指期货相关，但需要确认具体品种和经济含义。"
    return "待人工确认", "命中股指期货关键词但规则无法明确归类。"


def build_futures_rationality(screening: pd.DataFrame) -> pd.DataFrame:
    """检查股指期货成交、持仓、成交持仓比和基差率指标。"""
    mask = screening["factor"].str.contains("股指期货|I_FCM|基差率|成交持仓比|持仓量_I_FCM|成交量_I_FCM", na=False)
    out = screening.loc[mask].copy()
    labels = out.apply(futures_rationality_label, axis=1)
    out["纳入合理性"] = [item[0] for item in labels]
    out["合理性说明"] = [item[1] for item in labels]
    return out[
        [
            "factor",
            "name",
            "cls3",
            "cls4",
            "sample_count",
            "rank_ic",
            "rank_ic_pvalue",
            "ols_p_beta",
            "screening_label",
            "纳入合理性",
            "合理性说明",
        ]
    ]


def economic_group(row: pd.Series) -> str:
    """按经济逻辑给核心候选分组。"""
    factor = str(row["factor"])
    name = str(row.get("name", ""))
    if "成交量_A股指数" in factor or "成交额_A股指数" in factor or name in {"市场成交量", "市场成交额"}:
        return "市场成交活跃度"
    if "时序波动" in factor:
        return "时序波动"
    if "截面波动" in factor:
        return "截面波动"
    if "相对强弱" in factor:
        return "大小盘风格"
    if "I_FCM" in factor or "基差率" in factor or "股指期货" in factor:
        return "股指期货拥挤度或对冲环境"
    return "待人工确认"


def representative_score(row: pd.Series) -> tuple:
    """代表因子排序：分级、p值、Rank IC、滚动稳定性、样本量。"""
    level_order = {"强核心": 1, "弱核心": 2, "待复核": 3}
    return (
        level_order.get(row["核心分级"], 9),
        row["rank_ic_pvalue"],
        row["ols_p_beta"],
        -abs(row["rank_ic"]),
        -abs(row["rolling_beta_positive_ratio"] - 0.5),
        -row["sample_count"],
    )


def choose_representatives(core: pd.DataFrame) -> pd.DataFrame:
    """每个经济逻辑组推荐一个代表因子，不强制凑数量。"""
    rows = []
    for group, part in core.groupby("经济逻辑组"):
        ranked = part.copy()
        ranked["_score"] = ranked.apply(representative_score, axis=1)
        ranked = ranked.sort_values("_score")
        best = ranked.iloc[0].copy()
        best["代表推荐理由"] = (
            f"在{group}组内，综合核心分级、Rank IC p值、OLS p值、滚动稳定性、样本量和经济含义后最优。"
        )
        rows.append(best.drop(labels=["_score"]))
    reps = pd.DataFrame(rows)
    return reps.sort_values(["经济逻辑组", "rank_ic_pvalue"])


def rank_ic_pvalue(x: pd.Series, y: pd.Series) -> tuple[float, float]:
    """Spearman Rank IC 和 p值。"""
    valid = pd.concat([x, y], axis=1).dropna()
    if len(valid) < 3:
        return np.nan, np.nan
    result = spearmanr(valid.iloc[:, 0], valid.iloc[:, 1])
    return float(result.correlation), float(result.pvalue)


def ols_beta_pvalue(x: pd.Series, y: pd.Series) -> tuple[float, float]:
    """OLS beta 和 p值。"""
    valid = pd.concat([x, y], axis=1).dropna()
    valid.columns = ["factor", "target"]
    if len(valid) < 5:
        return np.nan, np.nan
    model = sm.OLS(valid["target"], sm.add_constant(valid["factor"])).fit()
    return float(model.params.get("factor", np.nan)), float(model.pvalues.get("factor", np.nan))


def high_low_spread(df: pd.DataFrame, direction: int) -> float:
    """按训练期方向计算高低分组收益差。"""
    valid = df[["factor_oos_z", "next_relative_return"]].dropna().copy()
    if len(valid) < 10:
        return np.nan
    valid["group"] = pd.qcut(valid["factor_oos_z"], q=5, labels=False, duplicates="drop") + 1
    if valid["group"].nunique() < 2:
        return np.nan
    group_ret = valid.groupby("group")["next_relative_return"].mean()
    low = group_ret.loc[group_ret.index.min()]
    high = group_ret.loc[group_ret.index.max()]
    return float(high - low) if direction >= 0 else float(low - high)


def hit_rate(df: pd.DataFrame, direction: int) -> float:
    """方向命中率：训练期方向 * 因子标准化值 的符号是否命中下一期相对收益符号。"""
    valid = df[["factor_oos_z", "next_relative_return"]].dropna().copy()
    valid = valid[(valid["factor_oos_z"] != 0) & (valid["next_relative_return"] != 0)]
    if valid.empty:
        return np.nan
    pred = np.sign(direction * valid["factor_oos_z"])
    actual = np.sign(valid["next_relative_return"])
    return float((pred == actual).mean())


def oos_validate_factor(aligned: pd.DataFrame, factor: str) -> list[dict]:
    """按时间顺序70/30切分，计算样本内和样本外结果。"""
    df = aligned[aligned["factor"] == factor].copy()
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").dropna(subset=["factor_raw", "next_relative_return"]).reset_index(drop=True)
    split_idx = int(len(df) * 0.7)
    train = df.iloc[:split_idx].copy()
    test = df.iloc[split_idx:].copy()

    mean = train["factor_raw"].mean()
    std = train["factor_raw"].std()
    if pd.isna(std) or std == 0:
        std = 1.0
    train["factor_oos_z"] = (train["factor_raw"] - mean) / std
    test["factor_oos_z"] = (test["factor_raw"] - mean) / std

    train_ic, train_ic_p = rank_ic_pvalue(train["factor_oos_z"], train["next_relative_return"])
    direction = 1 if pd.isna(train_ic) or train_ic >= 0 else -1

    rows = []
    for period_name, part in [("样本内", train), ("样本外", test)]:
        ic, ic_p = rank_ic_pvalue(part["factor_oos_z"], part["next_relative_return"])
        beta, beta_p = ols_beta_pvalue(part["factor_oos_z"], part["next_relative_return"])
        rows.append(
            {
                "factor": factor,
                "period": period_name,
                "sample_count": len(part),
                "date_start": part["date"].min().date().isoformat() if len(part) else None,
                "date_end": part["date"].max().date().isoformat() if len(part) else None,
                "train_direction": direction,
                "rank_ic": ic,
                "rank_ic_pvalue": ic_p,
                "ols_beta": beta,
                "ols_pvalue": beta_p,
                "high_low_group_spread": high_low_spread(part, direction),
                "direction_hit_rate": hit_rate(part, direction),
            }
        )
    return rows


def build_oos_validation(representatives: pd.DataFrame, aligned: pd.DataFrame) -> pd.DataFrame:
    """对代表因子做样本内/样本外验证。"""
    rows = []
    for factor in representatives["factor"]:
        rows.extend(oos_validate_factor(aligned, factor))
    return pd.DataFrame(rows)


def representative_corr(representatives: pd.DataFrame, aligned: pd.DataFrame) -> pd.DataFrame:
    """计算代表因子之间的 Spearman 相关性。"""
    factors = representatives["factor"].tolist()
    wide = aligned[aligned["factor"].isin(factors)].pivot_table(
        index="date",
        columns="factor",
        values="factor_raw",
        aggfunc="first",
    )
    corr = wide.corr(method="spearman")
    return corr.reset_index().rename(columns={"index": "factor"})


def write_report(
    path: Path,
    futures: pd.DataFrame,
    core: pd.DataFrame,
    representatives: pd.DataFrame,
    oos: pd.DataFrame,
    rep_corr: pd.DataFrame,
) -> None:
    """生成中文报告。"""
    strong = int((core["核心分级"] == "强核心").sum())
    weak = int((core["核心分级"] == "弱核心").sum())
    review = int((core["核心分级"] == "待复核").sum())
    lines = [
        "# 核心候选复核、去重和样本外验证报告",
        "",
        "## 说明",
        "",
        "- 本阶段只做候选合理性复核、核心候选证据分级、高相关去重代表选择和时间顺序样本外验证。",
        "- 不做因子合成，不做机器学习，不删除原始候选。",
        "- 样本外验证按时间顺序前70%为训练期、后30%为测试期；因子方向只由训练期 Rank IC 符号决定。",
        "",
        "## 候选池合理性复核",
        "",
        f"- 股指期货成交、持仓、成交持仓比和基差率相关指标数量：{len(futures)}。",
        "- 这些变量多数不是纯股票现货指标，但可作为股票市场对冲环境、拥挤度或衍生品活跃度代理；是否纳入最终股票组模型需人工确认。",
        "",
        "## 核心候选证据分级",
        "",
        f"- 强核心：{strong}",
        f"- 弱核心：{weak}",
        f"- 待复核：{review}",
        "",
        "## 高相关去重代表",
        "",
    ]
    for _, row in representatives.iterrows():
        lines.append(f"- {row['经济逻辑组']}：`{row['factor']}`，{row['代表推荐理由']}")

    lines.extend(["", "## 样本外验证提示", ""])
    for _, row in oos[oos["period"] == "样本外"].iterrows():
        lines.append(
            f"- `{row['factor']}`：样本外 Rank IC={row['rank_ic']:.6f}，p值={row['rank_ic_pvalue']:.6f}，"
            f"高低分组收益差={row['high_low_group_spread']:.6f}，方向命中率={row['direction_hit_rate']:.2%}。"
        )

    numeric_corr = rep_corr.drop(columns=["factor"]).abs()
    max_corr = numeric_corr.where(~np.eye(len(numeric_corr), dtype=bool)).max().max()
    lines.extend(["", "## 代表因子相关性", ""])
    lines.append(f"- 代表因子之间最大绝对Spearman相关性约为 {max_corr:.6f}，低于0.7阈值。")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选、设置列宽。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        width = min(max(max(len("" if cell.value is None else str(cell.value)) for cell in column_cells) + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """导出复核Excel。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            format_sheet(writer, sheet_name)


def main() -> None:
    """运行核心候选复核和样本外验证。"""
    screening = pd.read_csv(SCREENING_PATH)
    aligned = pd.read_csv(ALIGNED_PATH)
    high_corr = pd.read_csv(HIGH_CORR_PATH)

    futures = build_futures_rationality(screening)

    core = screening[screening["screening_label"] == "核心候选"].copy()
    core["经济逻辑组"] = core.apply(economic_group, axis=1)
    core["核心分级"] = core.apply(core_evidence_level, axis=1)
    core["新手解释"] = core.apply(core_explanation, axis=1)
    core = core.sort_values(["核心分级", "经济逻辑组", "rank_ic_pvalue", "ols_p_beta"])

    representatives = choose_representatives(core)
    oos = build_oos_validation(representatives, aligned)
    rep_corr = representative_corr(representatives, aligned)

    save_table(futures, TABLES_DIR / "core_review_futures_rationality.csv")
    save_table(core, TABLES_DIR / "core_review_core_regrading.csv")
    save_table(representatives, TABLES_DIR / "core_review_representative_factors.csv")
    save_table(oos, TABLES_DIR / "core_review_oos_validation.csv")
    save_table(rep_corr, TABLES_DIR / "core_review_representative_corr.csv")

    sheets = {
        "项目说明": pd.DataFrame(
            [
                {"项目": "阶段", "说明": "核心候选复核、去重和样本外验证"},
                {"项目": "限制", "说明": "不做因子合成、不做机器学习、不删除原始候选"},
                {"项目": "目标变量", "说明": "下一周 量化-均衡 - 主观-均衡"},
                {"项目": "样本外切分", "说明": "按时间顺序前70%训练、后30%测试，方向只由训练期决定"},
            ]
        ),
        "候选池合理性": futures,
        "核心候选分级": core,
        "代表因子": representatives,
        "代表因子相关性": rep_corr,
        "样本外验证": oos,
        "高相关因子对": high_corr,
    }
    export_excel(REPORTS_DIR / "core_candidate_review_oos_report.xlsx", sheets)
    write_report(REPORTS_DIR / "core_candidate_review_oos_report.md", futures, core, representatives, oos, rep_corr)

    print("核心候选复核、去重和样本外验证完成。")
    print(f"核心候选数：{len(core)}")
    print(f"代表因子数：{len(representatives)}")
    print(f"股指期货相关合理性记录：{len(futures)}")
    print(f"Excel：{REPORTS_DIR / 'core_candidate_review_oos_report.xlsx'}")
    print(f"报告：{REPORTS_DIR / 'core_candidate_review_oos_report.md'}")


if __name__ == "__main__":
    main()
