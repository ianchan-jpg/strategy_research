"""导出分类因子阶段性最终 Excel 报告。"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"


def yes_no(value: bool) -> str:
    """把布尔值转成中文。"""
    return "是" if bool(value) else "否"


def factor_explanation(row: pd.Series) -> str:
    """给全部因子结果增加一句新手解释。"""
    direction = "正向" if row["rank_ic"] >= 0 else "反向"
    significant = row["是否显著"] == "是"
    monotonic = row["是否单调"] == "是"
    recommended = row["是否推荐"] == "是"

    if recommended:
        return (
            f"该因子属于{row['category']}，与下一期相对收益呈{direction}关系，"
            f"统计显著性较好，当前可作为该类或高相关组的代表候选。"
        )
    if significant:
        return (
            f"该因子属于{row['category']}，Rank IC或回归有一定显著性，"
            f"但分组{'较单调' if monotonic else '不单调'}，需要继续复核稳定性。"
        )
    return (
        f"该因子属于{row['category']}，当前统计证据偏弱，"
        "暂时更适合作为观察项而不是核心候选。"
    )


def group_shape_explanation(row: pd.Series) -> str:
    """解释分组收益形态。"""
    return (
        f"第{int(row['group'])}组的下一期相对收益均值为{row['mean_next_relative_return']:.4%}，"
        f"整体形态判断为：{row['nonlinear_shape']}。"
    )


def build_all_factor_sheet(summary: pd.DataFrame, screening: pd.DataFrame, representatives: pd.DataFrame) -> pd.DataFrame:
    """整理“全部因子结果”工作表。"""
    rep_factors = set(representatives.loc[representatives["factor"] == representatives["recommended_representative"], "factor"])
    out = screening.copy()
    out["rank_ic_abs"] = out["rank_ic"].abs()
    out["是否显著"] = ((out["rank_ic_pvalue"] <= 0.1) | (out["ols_p_beta"] <= 0.1)).map(yes_no)
    out["是否单调"] = out["group_return_monotonic"].map(yes_no)
    out["是否推荐"] = out["factor"].isin(rep_factors).map(yes_no)
    out["新手解释"] = out.apply(factor_explanation, axis=1)
    out = out.sort_values(["category", "rank_ic_abs"], ascending=[True, False])
    return out


def build_representative_sheet(representatives: pd.DataFrame) -> pd.DataFrame:
    """整理“分类代表因子”工作表。"""
    out = representatives.copy()
    if out.empty:
        return out
    out["rank_ic_abs"] = out["rank_ic"].abs()
    out["是否显著"] = ((out["rank_ic_pvalue"] <= 0.1) | (out["ols_p_beta"] <= 0.1)).map(yes_no)
    out["是否单调"] = out["group_return_monotonic"].map(yes_no)
    out["是否推荐"] = (out["factor"] == out["recommended_representative"]).map(yes_no)
    out["新手解释"] = out.apply(factor_explanation, axis=1)
    return out.sort_values(["category", "rank_ic_abs"], ascending=[True, False])


def build_high_corr_sheet(high_corr: pd.DataFrame, representatives: pd.DataFrame) -> pd.DataFrame:
    """整理“高相关因子”工作表。"""
    out = high_corr.copy()
    rep_map = representatives.set_index("factor")["recommended_representative"].to_dict() if not representatives.empty else {}
    out["abs_corr"] = out["corr"].abs()
    out["代表指标"] = out["factor_1"].map(rep_map).fillna(out["factor_2"].map(rep_map))
    out["是否显著"] = "不适用"
    out["是否单调"] = "不适用"
    out["是否推荐"] = out["代表指标"].notna().map(yes_no)
    out["新手解释"] = out.apply(
        lambda row: f"这两个因子的Spearman相关性为{row['corr']:.3f}，信息重叠较高，建议优先查看代表指标：{row['代表指标']}。",
        axis=1,
    )
    return out.sort_values("abs_corr", ascending=False)


def build_group_shape_sheet(group_shape: pd.DataFrame) -> pd.DataFrame:
    """整理“分组收益形态”工作表。"""
    out = group_shape.copy()
    out["是否显著"] = "不适用"
    out["是否单调"] = out["nonlinear_shape"].isin(["单调递增", "单调递减"]).map(yes_no)
    out["是否推荐"] = "不适用"
    out["新手解释"] = out.apply(group_shape_explanation, axis=1)
    return out.sort_values(["factor", "group"])


def build_project_intro() -> pd.DataFrame:
    """项目说明工作表。"""
    rows = [
        {"项目": "研究目标", "说明": "使用周度权益因子预测下一期“量化-均衡 - 主观-均衡”。"},
        {"项目": "预测方向", "说明": "Factor(t) -> RelativeReturn(t+1)，避免使用未来因子。"},
        {"项目": "本阶段范围", "说明": "只做初步量化筛选、相关性检查和分组形态判断，暂不做因子合成。"},
        {"项目": "主要方法", "说明": "Rank IC、Rank IC p值、OLS、52周滚动回归、5组分组收益、Spearman相关性。"},
        {"项目": "输出解释", "说明": "“是否显著”以Rank IC p值或OLS beta p值小于等于0.1为初步标准。"},
        {"项目": "风险提示", "说明": "当前结果只代表历史统计关系，不代表因果关系或未来稳定收益。"},
    ]
    return pd.DataFrame(rows)


def build_conclusion_sheet(all_factors: pd.DataFrame, high_corr: pd.DataFrame, group_shape: pd.DataFrame) -> pd.DataFrame:
    """阶段性结论工作表。"""
    cross_section = all_factors[all_factors["factor"].str.contains("截面波动", na=False)]
    significant_cross_section = int((cross_section["是否显著"] == "是").sum())
    extreme_count = int(group_shape["nonlinear_shape"].astype(str).str.contains("极端", na=False).sum())
    rows = [
        {
            "结论": "初步筛选定位",
            "说明": "当前是初步量化筛选，用于发现候选因子和风险点，不代表因果关系。",
        },
        {
            "结论": "截面波动类表现",
            "说明": f"截面波动类因子整体表现相对较好，其中显著候选数量为{significant_cross_section}个，但同类因子之间存在较高相关性。",
        },
        {
            "结论": "非线性形态",
            "说明": f"部分因子只在极端分位有效，当前保留因子的分组形态中有{extreme_count}行体现极端组特征，不能简单理解为线性单调关系。",
        },
        {
            "结论": "高相关处理",
            "说明": f"发现{len(high_corr)}组高相关因子对，本报告只推荐代表指标，不自动删除任何原始因子。",
        },
        {
            "结论": "后续工作",
            "说明": "后续再考虑因子合成和样本外验证，并重点检查交易成本、参数稳定性和不同市场阶段表现。",
        },
    ]
    return pd.DataFrame(rows)


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选并美化表头。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    """生成 category_factor_final_report.xlsx。"""
    summary = pd.read_csv(TABLES_DIR / "category_factor_summary.csv")
    screening = pd.read_csv(TABLES_DIR / "category_factor_screening.csv")
    high_corr = pd.read_csv(TABLES_DIR / "category_factor_high_corr_pairs.csv")
    representatives = pd.read_csv(TABLES_DIR / "category_factor_representatives.csv")
    group_shape = pd.read_csv(TABLES_DIR / "retained_factor_group_shape_detail.csv")

    intro = build_project_intro()
    all_factors = build_all_factor_sheet(summary, screening, representatives)
    representative_sheet = build_representative_sheet(representatives)
    high_corr_sheet = build_high_corr_sheet(high_corr, representatives)
    group_shape_sheet = build_group_shape_sheet(group_shape)
    conclusion = build_conclusion_sheet(all_factors, high_corr, group_shape)

    output_path = REPORTS_DIR / "category_factor_final_report.xlsx"
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        intro.to_excel(writer, sheet_name="项目说明", index=False)
        all_factors.to_excel(writer, sheet_name="全部因子结果", index=False)
        representative_sheet.to_excel(writer, sheet_name="分类代表因子", index=False)
        high_corr_sheet.to_excel(writer, sheet_name="高相关因子", index=False)
        group_shape_sheet.to_excel(writer, sheet_name="分组收益形态", index=False)
        conclusion.to_excel(writer, sheet_name="阶段性结论", index=False)

        for sheet_name in ["项目说明", "全部因子结果", "分类代表因子", "高相关因子", "分组收益形态", "阶段性结论"]:
            format_sheet(writer, sheet_name)

    print(f"已生成：{output_path}")


if __name__ == "__main__":
    main()
