"""导出修正版分类因子最终报告。

本脚本只读取已有 CSV 结果，不重新计算任何统计指标。
修正重点：
- 统一筛选标签、是否推荐和新手解释；
- 明确 Rank IC / OLS 哪一项显著；
- 区分“统计证据”和“高相关组内代表”；
- 每个类别只给出一个分类代表因子。
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"


def cn_bool(value: bool) -> str:
    """布尔值转中文。"""
    return "是" if bool(value) else "否"


def significant_basis(row: pd.Series) -> str:
    """说明显著性来自 Rank IC、OLS，或二者均不显著。"""
    rank_sig = row["rank_ic_pvalue"] <= 0.1
    ols_sig = row["ols_p_beta"] <= 0.1
    if rank_sig and ols_sig:
        return "Rank IC和OLS均显著"
    if rank_sig:
        return "仅Rank IC显著"
    if ols_sig:
        return "仅OLS显著"
    return "Rank IC和OLS均不显著"


def is_stable_direction(row: pd.Series) -> bool:
    """滚动 beta 方向是否较稳定。"""
    ratio = row["rolling_beta_positive_ratio"]
    return ratio >= 0.6 or ratio <= 0.4


def evidence_level(row: pd.Series) -> str:
    """按用户指定口径生成证据等级。"""
    rank_sig = row["rank_ic_pvalue"] <= 0.1
    ols_sig = row["ols_p_beta"] <= 0.1
    group_rep = row.get("组内代表", "否") == "是"

    if rank_sig and ols_sig and is_stable_direction(row):
        return "核心候选"
    if (not rank_sig) and (not ols_sig) and group_rep:
        return "组内代表"
    if rank_sig or ols_sig or row.get("screening_label") == "待复核":
        return "弱候选"
    return "暂不保留"


def recommendation_from_label(label: str) -> str:
    """统一筛选标签和是否推荐。"""
    if label == "保留":
        return "是"
    if label == "待复核":
        return "待复核"
    return "否"


def factor_explanation(row: pd.Series) -> str:
    """生成不自相矛盾的新手解释。"""
    basis = row["显著依据"]
    level = row["证据等级"]
    direction = "正向" if row["rank_ic"] >= 0 else "反向"

    if level == "核心候选":
        return (
            f"该因子属于{row['category']}，与下一期相对收益呈{direction}关系，"
            f"{basis}，且52周滚动beta方向较稳定，可作为核心候选继续复核。"
        )
    if level == "弱候选":
        return (
            f"该因子属于{row['category']}，{basis}，但分组收益不单调或证据不完整，"
            "只能作为弱候选继续观察。"
        )
    if level == "组内代表":
        return (
            f"该因子属于{row['category']}，{basis}，不能称为有效因子；"
            "它只是高相关因子组内相对较好的代表，便于后续减少重复观察。"
        )
    return (
        f"该因子属于{row['category']}，{basis}，当前没有明显证据支持，暂不保留为候选。"
    )


def group_shape_explanation(row: pd.Series) -> str:
    """解释分组收益形态。"""
    return (
        f"第{int(row['group'])}组下一期相对收益均值为{row['mean_next_relative_return']:.4%}，"
        f"整体形态为：{row['nonlinear_shape']}。"
    )


def prepare_all_factors(screening: pd.DataFrame, representatives: pd.DataFrame) -> pd.DataFrame:
    """构造“全部因子结果”。"""
    rep_factors = set(representatives.loc[representatives["factor"] == representatives["recommended_representative"], "factor"])

    out = screening.copy()
    out["rank_ic_abs"] = out["rank_ic"].abs()
    out["是否显著"] = ((out["rank_ic_pvalue"] <= 0.1) | (out["ols_p_beta"] <= 0.1)).map(cn_bool)
    out["显著依据"] = out.apply(significant_basis, axis=1)
    out["是否单调"] = out["group_return_monotonic"].map(cn_bool)
    out["是否推荐"] = out["screening_label"].map(recommendation_from_label)
    out["组内代表"] = out["factor"].isin(rep_factors).map(cn_bool)
    out["证据等级"] = out.apply(evidence_level, axis=1)
    out["新手解释"] = out.apply(factor_explanation, axis=1)
    return out.sort_values(["category", "rank_ic_abs"], ascending=[True, False])


def choose_category_representatives(all_factors: pd.DataFrame) -> pd.DataFrame:
    """每个类别只选一个代表因子。"""
    level_order = {"核心候选": 1, "弱候选": 2, "组内代表": 3, "暂不保留": 4}
    rec_order = {"是": 1, "待复核": 2, "否": 3}
    tmp = all_factors.copy()
    tmp["level_order"] = tmp["证据等级"].map(level_order)
    tmp["rec_order"] = tmp["是否推荐"].map(rec_order)
    tmp = tmp.sort_values(
        ["category", "level_order", "rec_order", "rank_ic_abs", "sample_count"],
        ascending=[True, True, True, False, False],
    )
    picked = tmp.groupby("category", as_index=False).first()
    out = picked[
        [
            "category",
            "factor",
            "rank_ic",
            "rank_ic_pvalue",
            "ols_p_beta",
            "rolling_beta_positive_ratio",
            "证据等级",
            "显著依据",
            "是否推荐",
            "组内代表",
        ]
    ].rename(
        columns={
            "category": "类别",
            "factor": "代表因子",
            "rank_ic": "Rank IC",
            "rank_ic_pvalue": "Rank IC p值",
            "ols_p_beta": "OLS p值",
            "rolling_beta_positive_ratio": "滚动beta正比例",
        }
    )

    out["推荐理由"] = out.apply(
        lambda row: (
            f"{row['证据等级']}；{row['显著依据']}；"
            f"是否推荐={row['是否推荐']}；组内代表={row['组内代表']}。"
        ),
        axis=1,
    )
    return out


def prepare_high_corr(high_corr: pd.DataFrame, representatives: pd.DataFrame) -> pd.DataFrame:
    """构造“高相关因子”。"""
    out = high_corr.copy()
    rep_map = representatives.set_index("factor")["recommended_representative"].to_dict() if not representatives.empty else {}
    out["abs_corr"] = out["corr"].abs()
    out["代表指标"] = out["factor_1"].map(rep_map).fillna(out["factor_2"].map(rep_map))
    out["是否显著"] = "不适用"
    out["是否单调"] = "不适用"
    out["是否推荐"] = out["代表指标"].notna().map(cn_bool)
    out["新手解释"] = out.apply(
        lambda row: f"这两个因子的Spearman相关性为{row['corr']:.3f}，信息重叠较高；只建议参考代表指标：{row['代表指标']}，不自动删除另一列。",
        axis=1,
    )
    return out.sort_values("abs_corr", ascending=False)


def prepare_group_shape(group_shape: pd.DataFrame) -> pd.DataFrame:
    """构造“分组收益形态”。"""
    out = group_shape.copy()
    out["是否显著"] = "不适用"
    out["是否单调"] = out["nonlinear_shape"].isin(["单调递增", "单调递减"]).map(cn_bool)
    out["是否推荐"] = "不适用"
    out["新手解释"] = out.apply(group_shape_explanation, axis=1)
    return out.sort_values(["factor", "group"])


def project_intro() -> pd.DataFrame:
    """项目说明。"""
    return pd.DataFrame(
        [
            {"项目": "研究目标", "说明": "使用周度权益因子预测下一期“量化-均衡 - 主观-均衡”。"},
            {"项目": "预测方向", "说明": "Factor(t) -> RelativeReturn(t+1)，不使用未来因子。"},
            {"项目": "本次修正", "说明": "只修正Excel报告口径，不重新计算统计结果。"},
            {"项目": "是否显著", "说明": "Rank IC p值<=0.1 或 OLS p值<=0.1。"},
            {"项目": "是否推荐", "说明": "保留=是；待复核=待复核；重复或证据不足=否。"},
            {"项目": "组内代表", "说明": "仅表示在高相关组内相对更适合观察，不等同于统计显著或有效因子。"},
        ]
    )


def conclusion_sheet(all_factors: pd.DataFrame, high_corr: pd.DataFrame, group_shape: pd.DataFrame) -> pd.DataFrame:
    """阶段性结论。"""
    cross_section = all_factors[all_factors["factor"].str.contains("截面波动", na=False)]
    core_count = int((all_factors["证据等级"] == "核心候选").sum())
    group_rep_count = int((all_factors["证据等级"] == "组内代表").sum())
    extreme_count = int(group_shape["nonlinear_shape"].astype(str).str.contains("极端", na=False).sum())
    return pd.DataFrame(
        [
            {"结论": "初步筛选定位", "说明": "当前是初步量化筛选，用于发现候选因子和风险点，不代表因果关系。"},
            {
                "结论": "截面波动类表现",
                "说明": f"截面波动类表现相对较好，但同类指标相关性较高；当前核心候选数量为{core_count}个，需要后续复核。",
            },
            {"结论": "高相关代表", "说明": f"有{group_rep_count}个因子属于组内代表，组内代表不等于统计显著，只是便于减少重复观察。"},
            {"结论": "非线性形态", "说明": f"部分因子只在极端分位有效，当前分组形态中有{extreme_count}行体现极端组特征。"},
            {"结论": "高相关处理", "说明": f"发现{len(high_corr)}组高相关因子对，本报告只推荐代表指标，不自动删除任何原始因子。"},
            {"结论": "后续工作", "说明": "后续再考虑因子合成和样本外验证，并重点检查交易成本、参数稳定性和不同市场阶段表现。"},
        ]
    )


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选并设置样式。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    """生成修正版报告。"""
    summary = pd.read_csv(TABLES_DIR / "category_factor_summary.csv")
    screening = pd.read_csv(TABLES_DIR / "category_factor_screening.csv")
    high_corr = pd.read_csv(TABLES_DIR / "category_factor_high_corr_pairs.csv")
    representatives = pd.read_csv(TABLES_DIR / "category_factor_representatives.csv")
    group_shape = pd.read_csv(TABLES_DIR / "retained_factor_group_shape_detail.csv")

    intro = project_intro()
    all_factors = prepare_all_factors(screening, representatives)
    category_reps = choose_category_representatives(all_factors)
    high_corr_sheet = prepare_high_corr(high_corr, representatives)
    group_shape_sheet = prepare_group_shape(group_shape)
    conclusion = conclusion_sheet(all_factors, high_corr, group_shape)

    output_path = REPORTS_DIR / "category_factor_final_report_v2.xlsx"
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        intro.to_excel(writer, sheet_name="项目说明", index=False)
        all_factors.to_excel(writer, sheet_name="全部因子结果", index=False)
        category_reps.to_excel(writer, sheet_name="分类代表因子", index=False)
        high_corr_sheet.to_excel(writer, sheet_name="高相关因子", index=False)
        group_shape_sheet.to_excel(writer, sheet_name="分组收益形态", index=False)
        conclusion.to_excel(writer, sheet_name="阶段性结论", index=False)

        for sheet in ["项目说明", "全部因子结果", "分类代表因子", "高相关因子", "分组收益形态", "阶段性结论"]:
            format_sheet(writer, sheet)

    print(f"已生成：{output_path}")


if __name__ == "__main__":
    main()
