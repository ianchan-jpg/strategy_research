"""整理股票组全量因子复核 Excel。

只读取已有 CSV，不重新计算统计结果。
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
TABLES_DIR = PROJECT_DIR / "outputs" / "tables"
REPORTS_DIR = PROJECT_DIR / "outputs" / "reports"


def explain(row: pd.Series) -> str:
    """用一句中文解释筛选标签。"""
    label = row["筛选标签"]
    rank_sig = row["Rank IC p值"] <= 0.1
    ols_sig = row["OLS p值"] <= 0.1
    stable = row["滚动beta正比例"] >= 0.6 or row["滚动beta正比例"] <= 0.4

    if label == "核心候选":
        return "Rank IC和OLS均显著，且滚动beta方向较稳定，因此归为核心候选。"
    if label == "弱候选":
        if rank_sig and not ols_sig:
            return "Rank IC显著但OLS不显著，说明排序关系较强但线性回归证据不足，因此归为弱候选。"
        if ols_sig and not rank_sig:
            return "OLS显著但Rank IC不显著，说明线性关系有一定证据但排序稳定性不足，因此归为弱候选。"
        if stable:
            return "有部分统计证据且滚动方向相对稳定，但未同时满足核心候选标准，因此归为弱候选。"
        return "有部分统计证据，但稳定性或显著性不够完整，因此归为弱候选。"
    if label == "组内代表":
        return "自身统计显著性不强，但在高相关因子组内相对更适合作为观察代表，因此归为组内代表。"
    return "Rank IC、OLS或滚动稳定性证据不足，暂时归为证据不足。"


def format_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """冻结首行、开启筛选并设置列宽。"""
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    """生成 full_stock_factor_review.xlsx。"""
    screening = pd.read_csv(TABLES_DIR / "full_stock_factor_screening.csv")
    summary = pd.read_csv(TABLES_DIR / "full_stock_factor_summary.csv")

    summary_cols = [
        "factor",
        "sample_count",
        "rank_ic",
        "rank_ic_pvalue",
        "ols_beta",
        "ols_p_beta",
        "ols_r2",
        "rolling_beta_positive_ratio",
        "group_return_monotonic",
    ]
    merged = screening.merge(
        summary[summary_cols],
        on="factor",
        suffixes=("", "_summary"),
        how="left",
    )

    out = pd.DataFrame(
        {
            "因子名称": merged["factor"],
            "类别": merged["cls3"].fillna("") + "/" + merged["cls4"].fillna(""),
            "样本数": merged["sample_count"],
            "Rank IC": merged["rank_ic"],
            "Rank IC p值": merged["rank_ic_pvalue"],
            "OLS beta": merged["ols_beta"],
            "OLS p值": merged["ols_p_beta"],
            "R²": merged["ols_r2"],
            "滚动beta正比例": merged["rolling_beta_positive_ratio"],
            "分组单调性": merged["group_return_monotonic"].map(lambda value: "是" if bool(value) else "否"),
            "筛选标签": merged["screening_label"],
            "筛选原因": merged["screening_reason"],
        }
    )
    out["Rank IC绝对值"] = out["Rank IC"].abs()
    out["新手解释"] = out.apply(explain, axis=1)

    label_order = {"核心候选": 1, "弱候选": 2, "组内代表": 3, "证据不足": 4}
    out["_label_order"] = out["筛选标签"].map(label_order).fillna(99)
    out = out.sort_values(["_label_order", "Rank IC绝对值"], ascending=[True, False]).drop(columns=["_label_order"])

    output_path = REPORTS_DIR / "full_stock_factor_review.xlsx"
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="股票组因子复核", index=False)
        format_sheet(writer, "股票组因子复核")

    print(f"已生成：{output_path}")


if __name__ == "__main__":
    main()
